from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
import json
import os
import re
import unicodedata
from collections import defaultdict
from pathlib import Path

from django.conf import settings
from django.db import models, transaction
from django.db.models import Count, Q, Sum, Value
from django.db.models.functions import Coalesce
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.views.decorators.cache import never_cache
from django.views.decorators.csrf import ensure_csrf_cookie
from django.views.decorators.http import require_POST

from .models import (
    CacarolaMaquina,
    CacarolaProduto,
    CacarolaRegistro,
    Categoria,
    DiscoRazaoSocial,
    FiltroPedido,
    OrdemProducao,
    Produto,
    Relacao,
    RelacaoItem,
)


PCP_EXPORTS = {
    "mega_placas_ind": {
        "nome": "MEGA PLACAS IND",
    },
    "sami_industria": {
        "nome": "SAMI INDUSTRIA",
    },
    "cssb_curitiba": {
        "nome": "CSSB - CURITIBA",
    },
    "cssb_lapa": {
        "nome": "CSSB - LAPA",
        # Se não houver upload salvo em media, usa este arquivo (quando existir no disco).
        "source_path": r"h:\Meu Drive\PCP\EXP_PEDIDOS\PEDIDOS.xls",
    },
}

BASE_FILTRO_XLSX_PATH = r"h:\Meu Drive\Uriel\base filtro.xlsx"
BASE_DISCOS_XLS_PATH = r"h:\Meu Drive\Uriel\base filtro R SOCIAL.xlsx"
DATA_XLSX_PATH = r"h:\Meu Drive\Uriel\DATA.xlsx"


def _pcp_saved_file_abs_path(export_key: str) -> Path:
    root = getattr(settings, "MEDIA_ROOT", None)
    if not root:
        root = Path(settings.BASE_DIR) / "media"
    return Path(root) / "pcp_imports" / f"{export_key}_Pedidos.xls"


def _pcp_saved_name_meta_path(export_key: str) -> Path:
    return _pcp_saved_file_abs_path(export_key).with_suffix(".name.txt")


def _pcp_save_uploaded_xls(export_key: str, uploaded_name: str, data: bytes) -> str | None:
    dest = _pcp_saved_file_abs_path(export_key)
    try:
        dest.parent.mkdir(parents=True, exist_ok=True)
        tmp = dest.with_suffix(".tmp")
        tmp.write_bytes(data)
        tmp.replace(dest)
        try:
            _pcp_saved_name_meta_path(export_key).write_text((uploaded_name or "").strip(), encoding="utf-8")
        except OSError:
            pass
        return None
    except OSError as exc:
        return str(exc)


def _pcp_saved_uploaded_name(export_key: str) -> str:
    p = _pcp_saved_name_meta_path(export_key)
    if p.is_file():
        try:
            return (p.read_text(encoding="utf-8") or "").strip()
        except OSError:
            return ""
    return ""


def _pcp_effective_file_path(export_key: str) -> Path | None:
    """Arquivo a ler: upload em media tem prioridade; senão, source_path da fonte (se existir)."""
    saved = _pcp_saved_file_abs_path(export_key)
    if saved.is_file():
        return saved
    meta = PCP_EXPORTS.get(export_key) or {}
    ext = (meta.get("source_path") or "").strip()
    if ext:
        p = Path(ext)
        if p.is_file():
            return p
    return None


def _norm_header_txt(val) -> str:
    s = str(val or "").strip().lower()
    # normalização simples sem dependências extras
    repl = {
        "á": "a",
        "à": "a",
        "ã": "a",
        "â": "a",
        "é": "e",
        "ê": "e",
        "í": "i",
        "ó": "o",
        "ô": "o",
        "õ": "o",
        "ú": "u",
        "ç": "c",
    }
    for a, b in repl.items():
        s = s.replace(a, b)
    return re.sub(r"\s+", " ", s).strip()


def _find_pcp_col_indexes(headers: list[str]) -> dict[str, int | None]:
    h = [_norm_header_txt(x) for x in headers]

    def find_idx(pred):
        for i, col in enumerate(h):
            if pred(col):
                return i
        return None

    idx_pedido = find_idx(lambda c: "pedido" in c and ("numero" in c or "n " in c or "n." in c or "num" in c))
    if idx_pedido is None:
        idx_pedido = find_idx(lambda c: c == "pedido" or c.startswith("pedido "))
    idx_desc = find_idx(lambda c: "descricao" in c and ("produto" in c or "item" in c))
    if idx_desc is None:
        idx_desc = find_idx(lambda c: "descricao" in c)
    idx_cod = find_idx(lambda c: "cod" in c and "produto" in c)
    if idx_cod is None:
        idx_cod = find_idx(lambda c: c.startswith("cod") or "codigo" in c)
    idx_saldo = find_idx(lambda c: "saldo" in c and "pedido" in c)
    if idx_saldo is None:
        idx_saldo = find_idx(lambda c: "saldo" in c)
    idx_cliente = find_idx(lambda c: "cod" in c and "cliente" in c)
    if idx_cliente is None:
        idx_cliente = find_idx(lambda c: "cliente" in c and ("cod" in c or "codigo" in c))

    idx_data_entrada = find_idx(lambda c: "entrada" in c and ("data" in c or "dt" in c or "date" in c))
    if idx_data_entrada is None:
        idx_data_entrada = find_idx(lambda c: "data" in c and "entrada" in c)
    if idx_data_entrada is None:
        idx_data_entrada = find_idx(lambda c: "entrada" in c and "pedido" in c)
    if idx_data_entrada is None:
        idx_data_entrada = find_idx(lambda c: c == "entrada" or c.startswith("entrada "))

    return {
        "pedido": idx_pedido,
        "descricao": idx_desc,
        "codigo": idx_cod,
        "saldo": idx_saldo,
        "cod_cliente": idx_cliente,
        "data_entrada": idx_data_entrada,
    }


def _pcp_cell_to_date_iso(sh, wb, row_idx: int, col_idx: int | None) -> str:
    """Lê célula de planilha PCP como data e devolve YYYY-MM-DD ou ''."""
    if col_idx is None or col_idx < 0 or col_idx >= sh.ncols:
        return ""
    try:
        import xlrd
    except ImportError:
        return ""
    ctype = sh.cell_type(row_idx, col_idx)
    val = sh.cell_value(row_idx, col_idx)
    if ctype == xlrd.XL_CELL_DATE:
        try:
            dt = xlrd.xldate.xldate_as_datetime(float(val), wb.datemode)
            return dt.date().isoformat()
        except Exception:
            pass
    s = str(val or "").strip()
    if not s:
        return ""
    m = re.match(r"^(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{2,4})$", s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 100:
            y += 2000
        try:
            return date(y, mo, d).isoformat()
        except Exception:
            return ""
    m2 = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})$", s)
    if m2:
        y, mo, d = int(m2.group(1)), int(m2.group(2)), int(m2.group(3))
        try:
            return date(y, mo, d).isoformat()
        except Exception:
            return ""
    return ""


def _norm_code_key(val: str) -> str:
    raw = str(val or "").strip()
    if not raw:
        return ""
    digits = re.sub(r"\D", "", raw)
    if digits:
        try:
            return str(int(digits))
        except ValueError:
            return digits
    return re.sub(r"\s+", "", raw).upper()


_DATA_CUBAGEM_CACHE: dict[str, object] = {"path": None, "mtime": None, "map": {}}


def _to_decimal_cell(v) -> Decimal | None:
    if v is None:
        return None
    s = str(v).strip().replace(",", ".")
    if not s:
        return None
    try:
        return Decimal(s)
    except Exception:
        return None


def _load_data_cubagem_map(path: str) -> tuple[dict[str, tuple[Decimal | None, Decimal | None]], str | None]:
    """
    Lê DATA.xlsx (aba DATA) e retorna:
    codigo_normalizado -> (mv, m3_caixa)
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {}, "Instale openpyxl: pip install openpyxl"
    try:
        wb = load_workbook(filename=path, read_only=True, data_only=True)
        sh = wb["DATA"] if "DATA" in wb.sheetnames else wb[wb.sheetnames[0]]
    except Exception as exc:  # noqa: BLE001
        return {}, f"Não foi possível ler DATA.xlsx: {exc}"

    rows = list(sh.iter_rows(values_only=True))
    if not rows:
        return {}, "DATA.xlsx está vazio."

    headers = [_norm_header_txt(str(c or "").strip()) for c in rows[0]]

    def find_col(pred):
        for i, h in enumerate(headers):
            if pred(h):
                return i
        return None

    idx_codigo = find_col(
        lambda h: ("codigo" in h)
        or h == "cod"
        or h.startswith("cod ")
        or "cod " in h
        or "cod" in h
    )
    idx_mv = find_col(
        lambda h: h == "mv"
        or ("mult" in h and "venda" in h)
        or ("multiplo" in h and "venda" in h)
    )
    idx_m3 = find_col(
        lambda h: h == "m3"
        or ("cubagem" in h and "m3" in h)
        or ("cubagem" in h and "(m3)" in h)
    )
    if idx_codigo is None or idx_mv is None or idx_m3 is None:
        return {}, "Cabeçalhos esperados não encontrados na aba DATA (Código, MV, M3)."

    out: dict[str, tuple[Decimal | None, Decimal | None]] = {}
    for row in rows[1:]:
        codigo = str(row[idx_codigo] or "").strip() if idx_codigo < len(row) else ""
        if not codigo:
            continue
        k = _norm_code_key(codigo)
        mv = _to_decimal_cell(row[idx_mv] if idx_mv < len(row) else None)
        m3 = _to_decimal_cell(row[idx_m3] if idx_m3 < len(row) else None)
        out[k] = (mv, m3)
    return out, None


def _data_cubagem_map_cached() -> tuple[dict[str, tuple[Decimal | None, Decimal | None]], str | None]:
    path = DATA_XLSX_PATH
    if not os.path.isfile(path):
        return {}, f"Arquivo DATA.xlsx não encontrado em: {path}"
    try:
        mt = os.path.getmtime(path)
    except OSError as exc:
        return {}, str(exc)
    if (
        _DATA_CUBAGEM_CACHE.get("path") == path
        and _DATA_CUBAGEM_CACHE.get("mtime") == mt
        and isinstance(_DATA_CUBAGEM_CACHE.get("map"), dict)
    ):
        return _DATA_CUBAGEM_CACHE.get("map") or {}, None
    mp, err = _load_data_cubagem_map(path)
    if err:
        return {}, err
    _DATA_CUBAGEM_CACHE["path"] = path
    _DATA_CUBAGEM_CACHE["mtime"] = mt
    _DATA_CUBAGEM_CACHE["map"] = mp
    return mp, None


def _litros_para_m3_divisor() -> Decimal:
    """Litros (capacidade no texto) → m³: divide por este valor (settings.PRODUCAO_LITROS_PARA_M3_DIVISOR)."""
    raw = getattr(settings, "PRODUCAO_LITROS_PARA_M3_DIVISOR", 1000)
    try:
        d = Decimal(str(raw))
        if d > 0:
            return d
    except Exception:
        pass
    return Decimal("1000")


def _litros_para_m3_divisor_mercadao() -> Decimal:
    """Divisor para linhas MERC (panela de pressão do Mercadão)."""
    raw = getattr(settings, "PRODUCAO_LITROS_PARA_M3_DIVISOR_MERCADAO", 730)
    try:
        d = Decimal(str(raw))
        if d > 0:
            return d
    except Exception:
        pass
    return Decimal("730")


def _litros_para_m3_divisor_pratic_casa() -> Decimal:
    """Divisor para linhas PRATIC CASA (panela de pressao)."""
    raw = getattr(settings, "PRODUCAO_LITROS_PARA_M3_DIVISOR_PRATIC_CASA", 260.62)
    try:
        d = Decimal(str(raw))
        if d > 0:
            return d
    except Exception:
        pass
    return Decimal("260.62")


def _litros_divisor_efetivo(relacao_litros_divisor: Decimal | None = None) -> Decimal:
    """Usa sempre o divisor global configurado em settings."""
    _ = relacao_litros_divisor
    return _litros_para_m3_divisor()


def _ascii_upper(raw: str) -> str:
    s = str(raw or "")
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return s.upper()


def _extract_m3_from_text(raw: str) -> Decimal | None:
    """
    Extrai cubagem em m3/m³ de um texto livre.
    Ex.: "GAS ... 9,0 M3" -> 9.0
    Ignora padrões de densidade tipo "KG/M3".
    """
    s = str(raw or "")
    if not s:
        return None
    for m in re.finditer(r"(\d+(?:[.,]\d+)?)\s*m[³3]\b", s, flags=re.I):
        # Evita capturar "kg/m3" e variações com barra antes do m3.
        if m.start() > 0 and s[m.start() - 1] == "/":
            continue
        num_txt = (m.group(1) or "").replace(",", ".")
        try:
            val = Decimal(num_txt)
            if val > 0:
                return val
        except Exception:
            continue
    return None


def _litros_panela_pressao_mr_cook(desc: str) -> Decimal | None:
    """PANELA DE PRESSÃO + MR COOK X,X (capacidade em litros)."""
    s = str(desc or "")
    if not s:
        return None
    up = _ascii_upper(s)
    if "PANELA" not in up:
        return None
    if "PRESSAO" not in up:
        return None
    m = re.search(r"\bMR\s*COOK\s+(\d{1,2}(?:[.,]\d{1,2})?)\b", s, flags=re.I)
    if not m:
        return None
    try:
        litros = Decimal((m.group(1) or "").replace(",", "."))
        if litros <= 0:
            return None
        return litros
    except Exception:
        return None


def _litros_panela_pressao_mercadao(desc: str) -> Decimal | None:
    """PANELA DE PRESSAO MERC x,x (capacidade em litros)."""
    s = str(desc or "")
    if not s:
        return None
    up = _ascii_upper(s)
    if "PANELA" not in up or "PRESSAO" not in up or "MERC" not in up:
        return None
    m = re.search(r"\bMERC\b[^0-9]*(\d{1,2}(?:[.,]\d{1,2})?)\b", s, flags=re.I)
    if not m:
        return None
    try:
        litros = Decimal((m.group(1) or "").replace(",", "."))
        if litros <= 0:
            return None
        return litros
    except Exception:
        return None


def _litros_panela_pressao_mr_cook_45(desc: str) -> Decimal | None:
    """Fallback para Panela de Pressao MR COOK 4,5 quando o padrao principal nao casa."""
    s = str(desc or "")
    if not s:
        return None
    up = _ascii_upper(s)
    if "PANELA" not in up or "PRESSAO" not in up or "MR COOK" not in up:
        return None
    m = re.search(r"\b4[.,]5\b", s)
    if not m:
        return None
    return Decimal("4.5")


def _litros_panela_pressao_pratic_casa(desc: str) -> Decimal | None:
    """PANELA DE PRESSAO PRATIC CASA x,x (capacidade em litros)."""
    s = str(desc or "")
    if not s:
        return None
    up = _ascii_upper(s)
    if "PANELA" not in up or "PRESSAO" not in up or "PRATIC CASA" not in up:
        return None
    m = re.search(r"\bPRATIC\s+CASA\b[^0-9]*(\d{1,2}(?:[.,]\d{1,2})?)\b", s, flags=re.I)
    if not m:
        return None
    try:
        litros = Decimal((m.group(1) or "").replace(",", "."))
        if litros <= 0:
            return None
        return litros
    except Exception:
        return None


def _cubagem_m3_from_produto_descricao(
    desc: str, litros_divisor: Decimal | None = None
) -> Decimal | None:
    """
    m³ por unidade a partir da descrição (relação ou FiltroPedido).

    1) Valor explícito em m³ no texto (ex.: gás 9,0 M3).
    2) Litros inferidos (panela pressão ou Nxx) ÷ litros_divisor (relação) ou padrão em settings.
    """
    s = str(desc or "")
    div = _litros_divisor_efetivo(litros_divisor)
    v = _extract_m3_from_text(s)
    if v is not None:
        return v
    lit_pp = _litros_panela_pressao_mr_cook(s)
    if lit_pp is not None:
        return lit_pp / div
    lit_merc = _litros_panela_pressao_mercadao(s)
    if lit_merc is not None:
        return lit_merc / _litros_para_m3_divisor_mercadao()
    lit_pp_45 = _litros_panela_pressao_mr_cook_45(s)
    if lit_pp_45 is not None:
        return lit_pp_45 / _litros_para_m3_divisor_mercadao()
    lit_pratic = _litros_panela_pressao_pratic_casa(s)
    if lit_pratic is not None:
        return lit_pratic / _litros_para_m3_divisor_pratic_casa()
    up = _ascii_upper(s)
    if not s or "MR COOK" not in up:
        return None
    m = re.search(r"\bN(\d{1,2})\b", s, flags=re.I)
    if not m:
        return None
    try:
        n = int(m.group(1))
        if n >= 10:
            litros = Decimal(n) / Decimal(10)
        else:
            litros = Decimal(n)
        return litros / div
    except Exception:
        return None


def _cubagem_maps_from_filtro(
    codigos: set[str], litros_divisor: Decimal | None = None
) -> tuple[dict[str, Decimal], dict[str, Decimal]]:
    """
    Do FiltroPedido: volume_m3 explícito por código (override) e cubagem inferida do texto (fallback).
    Várias linhas com o mesmo código: usa a primeira descrição que gerar cubagem válida.
    """
    volume_m3_por_cod: dict[str, Decimal] = {}
    texto_m3_por_cod: dict[str, Decimal] = {}
    for fp in FiltroPedido.objects.exclude(cod_interno="").all():
        k = _norm_code_key(fp.cod_interno)
        if not k or k not in codigos:
            continue
        if k not in volume_m3_por_cod and fp.volume_m3 is not None:
            try:
                vm = Decimal(fp.volume_m3)
                if vm > 0:
                    volume_m3_por_cod[k] = vm
            except Exception:
                pass
        if k in texto_m3_por_cod:
            continue
        v = _cubagem_m3_from_produto_descricao(
            fp.descricao, litros_divisor
        ) or _cubagem_m3_from_produto_descricao(fp.descricao_produto, litros_divisor)
        if v is not None:
            texto_m3_por_cod[k] = v
    return volume_m3_por_cod, texto_m3_por_cod


def _cubagem_unit_para_item(
    k: str,
    volume_m3_por_cod: dict[str, Decimal],
    texto_m3_por_cod: dict[str, Decimal],
    descricao: str,
    litros_divisor: Decimal | None = None,
) -> Decimal | None:
    """
    1) volume_m3 no FiltroPedido (cadastro explícito);
    2) texto do FiltroPedido para o código (produto canônico na base);
    3) descrição da linha na relação (só se não houver cubagem pelo código).
    """
    if k and k in volume_m3_por_cod:
        return volume_m3_por_cod[k]
    if k and k in texto_m3_por_cod:
        return texto_m3_por_cod[k]
    return _cubagem_m3_from_produto_descricao(descricao, litros_divisor)


def _fonte_cubagem_item(
    k: str,
    volume_m3_por_cod: dict[str, Decimal],
    texto_m3_por_cod: dict[str, Decimal],
    descricao: str,
    litros_divisor: Decimal | None = None,
) -> str:
    """Rótulo coerente com a mesma ordem de _cubagem_unit_para_item."""
    if k and k in volume_m3_por_cod:
        return "cadastro (volume_m3)"
    if k and k in texto_m3_por_cod:
        return "base filtro (texto do produto)"
    if _cubagem_m3_from_produto_descricao(descricao, litros_divisor) is not None:
        return "descrição da linha"
    if k:
        return "sem cubagem — preencha volume_m3 em Admin → FiltroPedido"
    return "sem código de produto"


def _relacao_m3_linhas(itens_qs) -> tuple[Decimal, list[dict]]:
    """
    Total m³ e uma linha por item com cubagem, quantidade e subtotal (para conferência).
    """
    items = list(itens_qs)
    data_map, data_err = _data_cubagem_map_cached()

    total = Decimal("0")
    linhas: list[dict] = []
    for it in items:
        k = _norm_code_key(it.codigo_produto or "")
        desc = it.descricao or ""
        try:
            qtd = Decimal(it.quantidade or 0)
        except Exception:
            qtd = Decimal("0")

        sub: Decimal | None = None
        cub: Decimal | None = None
        fonte = ""
        if data_err:
            fonte = f"erro DATA.xlsx: {data_err}"
        elif not k:
            fonte = "erro: sem código de produto"
        elif k not in data_map:
            fonte = "erro: código não encontrado na aba DATA"
        else:
            mv, m3_caixa = data_map[k]
            if mv is None or mv <= 0:
                fonte = "erro: MV ausente ou zero na aba DATA"
            elif m3_caixa is None or m3_caixa < 0:
                fonte = "erro: M3 ausente/inválido na aba DATA"
            else:
                # Regra principal: m³ por CAIXA (não por peça).
                caixas = qtd / mv
                sub = caixas * m3_caixa
                cub = m3_caixa / mv  # equivalente por peça (apenas informativo)
                total += sub
                fonte = "DATA.xlsx (caixas: qtd/MV * M3)"
        linhas.append(
            {
                "item_id": it.id,
                "codigo": it.codigo_produto or "",
                "descricao": desc,
                "quantidade": qtd,
                "m3_unidade": cub,
                "m3_linha": sub,
                "fonte": fonte,
            }
        )
    return total, linhas


def _relacao_total_m3(itens_qs) -> Decimal:
    """
    Soma m³ = quantidade × cubagem por item (apenas itens do queryset passado).

    Cubagem vem de volume_m3 no cadastro, da descrição da linha ou do FiltroPedido,
    sempre via _cubagem_m3_from_produto_descricao (regra única).
    """
    total, _ = _relacao_m3_linhas(itens_qs)
    return total


def _load_pcp_rows_from_xls(path: str) -> tuple[list[dict], str | None]:
    try:
        import xlrd
    except ImportError:
        return [], "Instale xlrd: pip install xlrd"
    try:
        wb = xlrd.open_workbook(path)
        sh = wb.sheet_by_index(0)
    except Exception as exc:  # noqa: BLE001
        return [], f"Não foi possível ler o arquivo: {exc}"

    if sh.nrows == 0:
        return [], None

    header_row = 0
    col_map = _find_pcp_col_indexes([sh.cell_value(0, c) for c in range(sh.ncols)])
    if all(v is None for v in col_map.values()):
        # tentativa: alguma planilha vem com 1a linha de título
        if sh.nrows > 1:
            header_row = 1
            col_map = _find_pcp_col_indexes([sh.cell_value(1, c) for c in range(sh.ncols)])

    def v_at(row_idx: int, col_idx: int | None) -> str:
        if col_idx is None or col_idx < 0 or col_idx >= sh.ncols:
            return ""
        val = sh.cell_value(row_idx, col_idx)
        if isinstance(val, float) and val == int(val):
            return str(int(val))
        return str(val or "").strip()

    rows: list[dict] = []
    start = header_row + 1
    for r in range(start, sh.nrows):
        pedido = v_at(r, col_map["pedido"])
        descricao = v_at(r, col_map["descricao"])
        codigo = v_at(r, col_map["codigo"])
        saldo = v_at(r, col_map["saldo"])
        cod_cliente = v_at(r, col_map.get("cod_cliente"))
        data_entrada_iso = _pcp_cell_to_date_iso(sh, wb, r, col_map.get("data_entrada"))

        # fallback mínimo por posição, se cabeçalho não foi identificado
        if all(v is None for v in col_map.values()):
            pedido = pedido or v_at(r, 0)
            descricao = descricao or v_at(r, 1)
            codigo = codigo or v_at(r, 2)
            saldo = saldo or v_at(r, 3)
            cod_cliente = cod_cliente or v_at(r, 4)

        if not (pedido or descricao or codigo or saldo or cod_cliente):
            continue
        rows.append(
            {
                "pedido": pedido,
                "descricao": descricao,
                "codigo": codigo,
                "saldo": saldo,
                "cod_cliente": cod_cliente,
                "data_entrada_iso": data_entrada_iso,
            }
        )
    return rows, None


def _load_discos_seed_from_xls(path: str) -> tuple[list[dict], str | None]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return [], "Instale openpyxl: pip install openpyxl"
    try:
        wb = load_workbook(filename=path, read_only=True, data_only=True)
        sh = wb[wb.sheetnames[0]]
    except Exception as exc:  # noqa: BLE001
        return [], f"Não foi possível ler arquivo base de Discos: {exc}"

    rows = list(sh.iter_rows(values_only=True))
    if not rows:
        return [], None

    def as_txt(v) -> str:
        if v is None:
            return ""
        if isinstance(v, float) and v == int(v):
            return str(int(v))
        return str(v).strip()

    first_a = _norm_header_txt(as_txt(rows[0][0] if len(rows[0]) > 0 else ""))
    first_b = _norm_header_txt(as_txt(rows[0][1] if len(rows[0]) > 1 else ""))
    start_row = 1 if (("cod" in first_a or "codigo" in first_a) and ("razao" in first_b or "social" in first_b)) else 0

    out: list[dict] = []
    for row in rows[start_row:]:
        cod = as_txt(row[0] if len(row) > 0 else "")
        raz = as_txt(row[1] if len(row) > 1 else "")
        if not (cod or raz):
            continue
        out.append({"cod": cod, "razao_social": raz})
    return out, None


def _sync_discos_from_source_if_available() -> str | None:
    if not os.path.isfile(BASE_DISCOS_XLS_PATH):
        return None
    rows, err = _load_discos_seed_from_xls(BASE_DISCOS_XLS_PATH)
    if err:
        return err
    if not rows:
        return None
    # Espelha a fonte exatamente (linha a linha), sem colapsar por código.
    DiscoRazaoSocial.objects.filter(origem="Pedidos.xls").delete()
    novos = [
        DiscoRazaoSocial(
            cod=(r.get("cod") or "").strip(),
            razao_social=(r.get("razao_social") or "").strip(),
            origem="Pedidos.xls",
        )
        for r in rows
    ]
    DiscoRazaoSocial.objects.bulk_create(novos, batch_size=1000)
    return None


def _seed_discos_if_empty() -> str | None:
    if DiscoRazaoSocial.objects.exists():
        return None
    if not os.path.isfile(BASE_DISCOS_XLS_PATH):
        return "Arquivo base de Discos não encontrado para carga inicial."
    rows, err = _load_discos_seed_from_xls(BASE_DISCOS_XLS_PATH)
    if err:
        return err
    if not rows:
        return "Base de Discos não possui linhas válidas."
    objs = [
        DiscoRazaoSocial(
            cod=(r.get("cod") or "").strip(),
            razao_social=(r.get("razao_social") or "").strip(),
            origem="Pedidos.xls",
        )
        for r in rows
    ]
    DiscoRazaoSocial.objects.bulk_create(objs, batch_size=1000)
    return None


def _load_rows_from_base_filtro_xlsx(path: str) -> tuple[list[dict], str | None]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return [], "Instale openpyxl: pip install openpyxl"
    try:
        wb = load_workbook(filename=path, read_only=True, data_only=True)
        sh = wb[wb.sheetnames[0]]
    except Exception as exc:  # noqa: BLE001
        return [], f"Não foi possível ler base filtro.xlsx: {exc}"

    all_rows = list(sh.iter_rows(values_only=True))
    if not all_rows:
        return [], None

    def find_cols(headers: list[str]) -> dict[str, int | None]:
        h = [_norm_header_txt(x) for x in headers]
        def idx(pred):
            for i, c in enumerate(h):
                if pred(c):
                    return i
            return None
        return {
            "descricao": idx(lambda c: "descricao" in c),
            "cod_interno": idx(lambda c: ("cod" in c or "codigo" in c) and "intern" in c),
            "setor": idx(lambda c: "setor" in c),
            "necessidade": idx(lambda c: "necessidade" in c),
        }

    header_row = 0
    headers = [str(c or "").strip() for c in all_rows[0]]
    col_map = find_cols(headers)
    if all(v is None for v in col_map.values()) and len(all_rows) > 1:
        header_row = 1
        headers = [str(c or "").strip() for c in all_rows[1]]
        col_map = find_cols(headers)

    def v_at(row_vals: tuple, idx: int | None) -> str:
        if idx is None or idx < 0 or idx >= len(row_vals):
            return ""
        val = row_vals[idx]
        if isinstance(val, float) and val == int(val):
            return str(int(val))
        return str(val or "").strip()

    parsed: list[dict] = []
    for row_vals in all_rows[header_row + 1 :]:
        desc = v_at(row_vals, col_map["descricao"])
        cod_interno = v_at(row_vals, col_map["cod_interno"])
        setor = v_at(row_vals, col_map["setor"])
        necessidade = v_at(row_vals, col_map["necessidade"])
        if all(v is None for v in col_map.values()):
            desc = desc or v_at(row_vals, 0)
            cod_interno = cod_interno or v_at(row_vals, 1)
            setor = setor or v_at(row_vals, 2)
            necessidade = necessidade or v_at(row_vals, 3)
        if not (desc or cod_interno or setor or necessidade):
            continue
        parsed.append(
            {
                "descricao": desc,
                "cod_interno": cod_interno,
                "setor": setor,
                "necessidade": necessidade,
            }
        )
    return parsed, None


def _seed_filtro_pedido_if_empty() -> str | None:
    has_any = FiltroPedido.objects.exists()
    has_new_layout_data = FiltroPedido.objects.filter(
        models.Q(descricao__gt="") | models.Q(cod_interno__gt="") | models.Q(setor__gt="") | models.Q(necessidade__gt="")
    ).exists()
    if has_any and has_new_layout_data:
        return None
    if not os.path.isfile(BASE_FILTRO_XLSX_PATH):
        return "Arquivo base filtro.xlsx não encontrado para carga inicial."
    rows, err = _load_rows_from_base_filtro_xlsx(BASE_FILTRO_XLSX_PATH)
    if err:
        return err
    if not rows:
        return "base filtro.xlsx não possui linhas válidas."
    # Recarrega a base quando ainda não está no layout novo.
    FiltroPedido.objects.all().delete()
    objs = [
        FiltroPedido(
            descricao=r["descricao"],
            cod_interno=r["cod_interno"],
            setor=r["setor"],
            necessidade=r["necessidade"],
            fonte="base filtro.xlsx",
            # mantém também campos legados preenchidos para compatibilidade
            descricao_produto=r["descricao"],
            codigo_produto=r["cod_interno"],
            saldo_pedido=r["necessidade"],
        )
        for r in rows
    ]
    FiltroPedido.objects.bulk_create(objs, batch_size=1000)
    return None


def home(request):
    ordens = OrdemProducao.objects.select_related("produto").order_by("-data_criacao")[:10]
    produtos = Produto.objects.filter(ativo=True).order_by("nome")[:10]
    contexto = {
        "ordens": ordens,
        "produtos": produtos,
    }
    return render(request, "producao/home.html", contexto)




def mega(request):
    return render(request, "producao/mega.html")


def ordem_servico_mega(request):
    return render(request, "producao/ordem_servico_mega.html")


@ensure_csrf_cookie
def pcp(request):
    erro_upload = None
    erro_gerar_relacao = None
    if request.method == "POST":
        action = (request.POST.get("acao") or "").strip()
        if action == "upload_pedidos":
            export_key = (request.POST.get("export_key") or "").strip()
            if export_key not in PCP_EXPORTS:
                erro_upload = "Fonte inválida para importação."
            else:
                file_obj = request.FILES.get("pedidos_file")
                if not file_obj:
                    erro_upload = "Selecione um arquivo .xls para importar."
                else:
                    filename = (file_obj.name or "").strip()
                    if not filename.lower().endswith(".xls"):
                        erro_upload = "Formato não suportado. Envie arquivo .xls."
                    else:
                        err = _pcp_save_uploaded_xls(export_key, filename, file_obj.read())
                        if err:
                            erro_upload = f"Não foi possível salvar o arquivo: {err}"
                        else:
                            ret = (request.POST.get("return_to") or "").strip()
                            if ret == "operacoes_importacao":
                                return redirect(f"{reverse('operacoes_importacao')}?saved=1")
                            return redirect(f"{reverse('pcp')}?saved=1")
        if action == "gerar_relacao":
            nome_relacao = (request.POST.get("nome_relacao") or "").strip()
            payload_raw = (request.POST.get("rows_payload") or "").strip()
            if not nome_relacao:
                erro_gerar_relacao = "Informe o nome da relação."
            elif not payload_raw:
                erro_gerar_relacao = "Nenhuma linha foi selecionada para gerar a relação."
            else:
                try:
                    payload = json.loads(payload_raw)
                except Exception:
                    payload = []
                if not isinstance(payload, list) or not payload:
                    erro_gerar_relacao = "Payload inválido para gerar relação."
                else:
                    try:
                        with transaction.atomic():
                            relacao = Relacao.objects.create(nome=nome_relacao)
                            objs: list[RelacaoItem] = []
                            for i, r in enumerate(payload):
                                codigo = str((r.get("codigo") or "")).strip()
                                descricao = str((r.get("descricao") or "")).strip()
                                pedido = str((r.get("pedido") or "")).strip()
                                odf = str((r.get("odf") or "")).strip()
                                qtd_raw = str((r.get("saldo") or "")).strip()
                                qtd = _parse_decimal(qtd_raw)
                                if not (codigo or descricao or qtd):
                                    continue
                                if codigo:
                                    Produto.objects.get_or_create(
                                        codigo=codigo,
                                        defaults={"nome": descricao or codigo},
                                    )
                                objs.append(
                                    RelacaoItem(
                                        relacao=relacao,
                                        indice=i + 1,
                                        descricao=descricao,
                                        codigo_produto=codigo,
                                        ok=False,
                                        parcial=Decimal("0"),
                                        quantidade=Decimal(str(qtd)),
                                        odf=odf,
                                        pedido_numero=pedido,
                                    )
                                )
                            if not objs:
                                raise ValueError("Nenhuma linha válida para criar relação.")
                            RelacaoItem.objects.bulk_create(objs, batch_size=500)
                        return redirect(
                            f"{reverse('relacao_detalhe', kwargs={'relacao_id': relacao.id})}?estoque_alerta=1"
                        )
                    except Exception as exc:  # noqa: BLE001
                        erro_gerar_relacao = f"Não foi possível gerar a relação: {exc}"

    exports = []
    for key, item in PCP_EXPORTS.items():
        saved = _pcp_saved_file_abs_path(key)
        uploaded_name = _pcp_saved_uploaded_name(key)
        effective = _pcp_effective_file_path(key)
        exports.append(
            {
                "key": key,
                "nome": item["nome"],
                "exists": effective is not None,
                "uploaded_name": uploaded_name,
            }
        )

    rows: list[dict] = []
    erros_tabela: list[str] = []
    fontes_disponiveis: list[dict] = []
    for key, item in PCP_EXPORTS.items():
        p = _pcp_effective_file_path(key)
        if not p:
            continue
        fonte_rows, err = _load_pcp_rows_from_xls(str(p))
        if err:
            erros_tabela.append(f"{item['nome']}: {err}")
            continue
        for r in fonte_rows:
            r["fonte_key"] = key
            r["fonte_nome"] = item["nome"]
        rows.extend(fonte_rows)
        fontes_disponiveis.append({"key": key, "nome": item["nome"]})

    # Fallback online: quando não há arquivos .xls disponíveis no servidor,
    # usa os dados persistidos em FiltroPedido para manter a aba PCP funcional.
    if not rows:
        for fp in FiltroPedido.objects.all().order_by("numero_pedido", "id"):
            pedido_fallback = (
                (fp.numero_pedido or "").strip()
                or (fp.codigo_produto or "").strip()
                or (fp.cod_interno or "").strip()
                or f"SEM-PEDIDO-{fp.id}"
            )
            rows.append(
                {
                    "pedido": pedido_fallback,
                    "descricao": (fp.descricao or fp.descricao_produto or "").strip(),
                    "codigo": (fp.cod_interno or fp.codigo_produto or "").strip(),
                    "saldo": (fp.necessidade or fp.saldo_pedido or "").strip(),
                    "fonte_key": "dados_filtragem",
                    "fonte_nome": "Dados de filtragem",
                    "cod_cliente": "",
                    "cliente": "",
                    "data_entrada_iso": "",
                }
            )
        if rows:
            fontes_disponiveis.append({"key": "dados_filtragem", "nome": "Dados de filtragem"})

    # Mapa Cód Interno -> Setor(es) vindo de "Dados de filtragem"
    setor_by_cod: dict[str, set[str]] = {}
    for f in FiltroPedido.objects.exclude(cod_interno="").exclude(setor="").all():
        k = _norm_code_key(f.cod_interno)
        if not k:
            continue
        setor_by_cod.setdefault(k, set()).add((f.setor or "").strip())

    for r in rows:
        k = _norm_code_key(r.get("codigo") or "")
        setores = sorted(setor_by_cod.get(k, set()))
        if not setores:
            setores = ["Sem setor"]
        r["setores_match"] = setores
        r["setores_match_str"] = "|".join(setores)

    related_pedidos = set(
        RelacaoItem.objects.exclude(pedido_numero="")
        .exclude(pedido_numero__isnull=True)
        .values_list("pedido_numero", flat=True)
        .distinct()
    )

    razao_map = {(_norm_code_key(x.cod)): (x.razao_social or "").strip() for x in DiscoRazaoSocial.objects.all()}
    for r in rows:
        k_cliente = _norm_code_key(r.get("cod_cliente") or "")
        k_prod = _norm_code_key(r.get("codigo") or "")
        raz = razao_map.get(k_cliente) or razao_map.get(k_prod) or "Sem razão social"
        r["razao_social"] = raz

    pedido_rel_nomes: dict[str, set[str]] = defaultdict(set)
    for item in RelacaoItem.objects.exclude(pedido_numero="").select_related("relacao"):
        pn = str(item.pedido_numero or "").strip()
        if not pn:
            continue
        nm = (item.relacao.nome or "").strip()
        if nm:
            pedido_rel_nomes[pn].add(nm)

    related_set = {str(x).strip() for x in related_pedidos if str(x or "").strip()}
    unique_pedidos = sorted({(r.get("pedido") or "").strip() for r in rows if (r.get("pedido") or "").strip()})
    pedido_setores: dict[str, set[str]] = defaultdict(set)
    for r in rows:
        p = (r.get("pedido") or "").strip()
        if not p:
            continue
        for s in r.get("setores_match") or []:
            ss = str(s or "").strip()
            if ss:
                pedido_setores[p].add(ss)
    pedido_razao: dict[str, str] = {}
    for r in rows:
        p = (r.get("pedido") or "").strip()
        if not p:
            continue
        rz = (r.get("razao_social") or "").strip() or "Sem razão social"
        if p not in pedido_razao:
            pedido_razao[p] = rz
        elif pedido_razao[p] in ("", "Sem razão social") and rz not in ("", "Sem razão social"):
            pedido_razao[p] = rz

    pedido_data_entrada: dict[str, str] = {}
    for r in rows:
        p = (r.get("pedido") or "").strip()
        d_iso = (r.get("data_entrada_iso") or "").strip()
        if not p or not d_iso:
            continue
        if p not in pedido_data_entrada or d_iso < pedido_data_entrada[p]:
            pedido_data_entrada[p] = d_iso

    pcp_pedidos_status: list[dict] = []
    for pedido in unique_pedidos:
        tirado = pedido in related_set
        rel_list = sorted(pedido_rel_nomes.get(pedido, set()))
        razao = pedido_razao.get(pedido, "Sem razão social")
        d_ent = pedido_data_entrada.get(pedido)
        pcp_pedidos_status.append(
            {
                "pedido": pedido,
                "tirado": tirado,
                "relacoes": rel_list,
                "razao_social": razao,
                "setores": sorted(pedido_setores.get(pedido, set())),
                "data_entrada_iso": d_ent,
            }
        )

    pedidos = sorted({(r.get("pedido") or "").strip() for r in rows if (r.get("pedido") or "").strip()})
    setores_disponiveis = sorted({s for r in rows for s in (r.get("setores_match") or []) if s})
    contexto = {
        "rows": rows,
        "pedidos": pedidos,
        "setores_disponiveis": setores_disponiveis,
        "related_pedidos": sorted({str(x).strip() for x in related_pedidos if str(x or "").strip()}),
        "fontes_disponiveis": fontes_disponiveis,
        "erros_tabela": erros_tabela,
        "erro_upload": erro_upload,
        "erro_gerar_relacao": erro_gerar_relacao,
        "upload_saved_notice": (request.GET.get("saved") or "").strip() == "1",
        "pcp_pedidos_status": pcp_pedidos_status,
    }
    if (
        erro_upload
        and request.method == "POST"
        and (request.POST.get("return_to") or "").strip() == "operacoes_importacao"
    ):
        return render(
            request,
            "producao/operacoes_importacao.html",
            {
                "exports": exports,
                "erro_upload": erro_upload,
                "upload_saved_notice": False,
            },
        )
    return render(request, "producao/pcp.html", contexto)


@ensure_csrf_cookie
def operacoes_importacao(request):
    """Página de importações do PCP (Pedidos.xls por fonte, MRP, estoque)."""
    exports = []
    for key, item in PCP_EXPORTS.items():
        uploaded_name = _pcp_saved_uploaded_name(key)
        effective = _pcp_effective_file_path(key)
        exports.append(
            {
                "key": key,
                "nome": item["nome"],
                "exists": effective is not None,
                "uploaded_name": uploaded_name,
            }
        )
    return render(
        request,
        "producao/operacoes_importacao.html",
        {
            "exports": exports,
            "erro_upload": None,
            "upload_saved_notice": (request.GET.get("saved") or "").strip() == "1",
        },
    )


def _build_discos_base_context() -> dict:
    # usa as mesmas fontes de pedidos já importadas no PCP
    rows: list[dict] = []
    for key, item in PCP_EXPORTS.items():
        p = _pcp_effective_file_path(key)
        if not p:
            continue
        fonte_rows, err = _load_pcp_rows_from_xls(str(p))
        if err:
            continue
        for r in fonte_rows:
            r["fonte_key"] = key
            r["fonte_nome"] = item["nome"]
        rows.extend(fonte_rows)

    razao_map = {(_norm_code_key(x.cod)): (x.razao_social or "").strip() for x in DiscoRazaoSocial.objects.all()}
    for r in rows:
        # Regra principal: Cód. Cliente -> Cód. (A) da base de razões.
        k_cliente = _norm_code_key(r.get("cod_cliente") or "")
        # Fallback: quando o arquivo não vier com Cód. Cliente, tenta pelo Cód. do Produto.
        k_prod = _norm_code_key(r.get("codigo") or "")
        raz = razao_map.get(k_cliente) or razao_map.get(k_prod) or "Sem razão social"
        r["razao_social"] = raz

    # Setor via "Dados de filtragem" (Cód. do Produto ↔ Cód Interno), igual à aba PCP
    setor_by_cod: dict[str, set[str]] = {}
    for f in FiltroPedido.objects.exclude(cod_interno="").exclude(setor="").all():
        k = _norm_code_key(f.cod_interno)
        if not k:
            continue
        setor_by_cod.setdefault(k, set()).add((f.setor or "").strip())

    for r in rows:
        k = _norm_code_key(r.get("codigo") or "")
        setores = sorted(setor_by_cod.get(k, set()))
        if not setores:
            setores = ["Sem setor"]
        r["setores_match"] = setores
        r["setores_match_str"] = "|".join(setores)

    # Filtro lateral: só razões que aparecem em linha com pedido; códigos (Cód. Cliente) para busca
    pedidos = sorted({(r.get("pedido") or "").strip() for r in rows if (r.get("pedido") or "").strip()})
    razao_to_cods: dict[str, set[str]] = defaultdict(set)
    for r in rows:
        if not (r.get("pedido") or "").strip():
            continue
        rz = (r.get("razao_social") or "").strip()
        if not rz:
            continue
        cc = (r.get("cod_cliente") or "").strip()
        if cc:
            razao_to_cods[rz].add(cc)
    razoes_busca = []
    for rz in sorted(razao_to_cods.keys()):
        tokens: set[str] = set()
        for c in razao_to_cods[rz]:
            raw = (c or "").strip()
            if raw:
                tokens.add(raw.lower())
            nk = _norm_code_key(c)
            if nk:
                tokens.add(nk.lower())
        razoes_busca.append({"razao": rz, "cods_busca": " ".join(sorted(tokens))})
    fontes = sorted({(r.get("fonte_key") or "").strip() for r in rows if (r.get("fonte_key") or "").strip()})
    setores_disponiveis = sorted({s for r in rows for s in (r.get("setores_match") or []) if s})

    return {
        "rows": rows,
        "razoes_busca": razoes_busca,
        "pedidos": pedidos,
        "fontes": fontes,
        "setores_disponiveis": setores_disponiveis,
    }


def discos(request):
    contexto = _build_discos_base_context()
    return render(request, "producao/discos.html", contexto)


def discos_pedidos(request):
    contexto = _build_discos_base_context()
    contexto["hoje"] = datetime.now()
    return render(request, "producao/discos_pedidos.html", contexto)


def dados_discos(request):
    erro_seed = _seed_discos_if_empty()
    if request.method == "POST":
        acao = (request.POST.get("acao") or "").strip()
        if acao == "salvar":
            itens = list(DiscoRazaoSocial.objects.all().order_by("id"))
            for it in itens:
                pre = f"row_{it.id}_"
                it.cod = (request.POST.get(pre + "cod") or "").strip()
                it.razao_social = (request.POST.get(pre + "razao_social") or "").strip()
                it.origem = (request.POST.get(pre + "origem") or "").strip()
                it.save()
            return redirect(f"{reverse('dados_discos')}?saved=1")
        if acao == "adicionar":
            cod = (request.POST.get("novo_cod") or "").strip()
            DiscoRazaoSocial.objects.create(
                cod=cod,
                razao_social=(request.POST.get("novo_razao_social") or "").strip(),
                origem=(request.POST.get("novo_origem") or "").strip() or "manual",
            )
            return redirect(f"{reverse('dados_discos')}?added=1")

    rows = list(DiscoRazaoSocial.objects.all().order_by("id"))
    return render(
        request,
        "producao/dados_discos.html",
        {
            "rows": rows,
            "erro_seed": erro_seed,
            "saved_notice": (request.GET.get("saved") or "").strip() == "1",
            "added_notice": (request.GET.get("added") or "").strip() == "1",
            "base_seed_path": BASE_DISCOS_XLS_PATH,
        },
    )


def dados_filtragem(request):
    erro = _seed_filtro_pedido_if_empty()

    if request.method == "POST":
        acao = (request.POST.get("acao") or "").strip()
        if acao == "salvar_linhas":
            itens = list(FiltroPedido.objects.all().order_by("numero_pedido", "id"))
            for it in itens:
                pre = f"row_{it.id}_"
                it.descricao = (request.POST.get(pre + "descricao") or "").strip()
                it.cod_interno = (request.POST.get(pre + "cod_interno") or "").strip()
                it.setor = (request.POST.get(pre + "setor") or "").strip()
                it.necessidade = (request.POST.get(pre + "necessidade") or "").strip()
                it.fonte = (request.POST.get(pre + "fonte") or "").strip()
                # espelha campos legados para não quebrar leituras anteriores
                it.descricao_produto = it.descricao
                it.codigo_produto = it.cod_interno
                it.saldo_pedido = it.necessidade
                it.save()
            return redirect(f"{reverse('dados_filtragem')}?saved=1")
        if acao == "adicionar_linha":
            FiltroPedido.objects.create(
                descricao=(request.POST.get("novo_descricao") or "").strip(),
                cod_interno=(request.POST.get("novo_cod_interno") or "").strip(),
                setor=(request.POST.get("novo_setor") or "").strip(),
                necessidade=(request.POST.get("novo_necessidade") or "").strip(),
                fonte=(request.POST.get("novo_fonte") or "").strip() or "manual",
                descricao_produto=(request.POST.get("novo_descricao") or "").strip(),
                codigo_produto=(request.POST.get("novo_cod_interno") or "").strip(),
                saldo_pedido=(request.POST.get("novo_necessidade") or "").strip(),
            )
            return redirect(f"{reverse('dados_filtragem')}?added=1")

    rows = list(FiltroPedido.objects.all().order_by("numero_pedido", "id"))
    return render(
        request,
        "producao/dados_filtragem.html",
        {
            "rows": rows,
            "erro_seed": erro,
            "saved_notice": (request.GET.get("saved") or "").strip() == "1",
            "added_notice": (request.GET.get("added") or "").strip() == "1",
            "base_seed_path": BASE_FILTRO_XLSX_PATH,
        },
    )

def relacoes(request):
    erro = None

    if request.method == "POST":
        nome = (request.POST.get("nome") or "").strip()
        colagem = (request.POST.get("colagem") or "").strip()

        if not nome:
            erro = "Informe o nome da relação."
        elif not colagem:
            erro = "Cole a tabela do Excel para importar."
        else:
            try:
                with transaction.atomic():
                    relacao = Relacao.objects.create(nome=nome)
                    itens = _parse_colagem_excel(colagem)
                    objetos_itens: list[RelacaoItem] = []
                    for i, item in enumerate(itens):
                        # garantir que cada código de produto vire um Produto único
                        codigo = (item["codigo_produto"] or "").strip()
                        if codigo:
                            Produto.objects.get_or_create(
                                codigo=codigo,
                                defaults={"nome": item["descricao"] or codigo},
                            )
                        objetos_itens.append(
                            RelacaoItem(
                                relacao=relacao,
                                indice=i + 1,
                                descricao=item["descricao"],
                                codigo_produto=item["codigo_produto"],
                                ok=item["ok"],
                                parcial=item.get("parcial", 0) or 0,
                                quantidade=item["quantidade"],
                                odf=item["odf"],
                                data=item["data"],
                            )
                        )
                    RelacaoItem.objects.bulk_create(objetos_itens)
            except Exception as exc:  # noqa: BLE001 - UX-friendly error
                erro = f"Não foi possível importar: {exc}"

        if not erro:
            return redirect(f"{reverse('relacoes')}?estoque_alerta=1")

    # filtros
    ok_filtro = (request.GET.get("ok") or "tudo").strip().lower()
    q = (request.GET.get("q") or "").strip()
    de = (request.GET.get("de") or "").strip()
    ate = (request.GET.get("ate") or "").strip()
    aba = (request.GET.get("aba") or "ativas").strip().lower()

    relacoes_qs = Relacao.objects.all().order_by("-criada_em")
    if aba == "finalizadas":
        relacoes_qs = relacoes_qs.filter(finalizada=True)
    else:
        relacoes_qs = relacoes_qs.filter(finalizada=False)

    # Para filtros que dependem de itens (OK/data/busca), vamos filtrar no nível de itens
    # e reduzir as relações pela existência de itens que casem.
    itens_qs = RelacaoItem.objects.all()

    if ok_filtro == "ok":
        itens_qs = itens_qs.filter(ok=True)
    elif ok_filtro == "pendente":
        itens_qs = itens_qs.filter(ok=False)

    if de:
        data_de = _parse_date_yyyy_mm_dd(de)
        if data_de:
            itens_qs = itens_qs.filter(data__gte=data_de)
    if ate:
        data_ate = _parse_date_yyyy_mm_dd(ate)
        if data_ate:
            itens_qs = itens_qs.filter(data__lte=data_ate)

    if q:
        itens_qs = itens_qs.filter(
            (
                models.Q(descricao__icontains=q)
                | models.Q(codigo_produto__icontains=q)
                | models.Q(odf__icontains=q)
            )
        )

    if ok_filtro != "tudo" or de or ate or q:
        relacoes_qs = relacoes_qs.filter(id__in=itens_qs.values_list("relacao_id", flat=True).distinct())

    contexto = {
        "erro": erro,
        "relacoes": relacoes_qs[:200],
        "aba": aba,
        "ok_filtro": ok_filtro,
        "q": q,
        "de": de,
        "ate": ate,
        "relacao_ativa_id": None,
    }
    return render(request, "producao/relacoes.html", contexto)


def relacao_detalhe(request, relacao_id: int):
    relacao = get_object_or_404(Relacao, pk=relacao_id)
    itens_base = relacao.itens.all().order_by("descricao", "codigo_produto", "id")

    if request.method == "POST":
        acao = request.POST.get("acao")
        if acao == "salvar":
            view_mode = (request.POST.get("view_mode") or "").strip().lower()
            if view_mode == "consolidada":
                itens_base_list = list(itens_base)
                grupos: dict[tuple[str, str], list[RelacaoItem]] = defaultdict(list)
                for it in itens_base_list:
                    chave = ((it.codigo_produto or "").strip(), (it.descricao or "").strip())
                    grupos[chave].append(it)
                grupos_ordenados = sorted(
                    list(grupos.items()),
                    key=lambda kv: ((kv[0][1] or "").lower(), (kv[0][0] or "").lower()),
                )
                grupos_por_idx = {str(i + 1): items for i, (_, items) in enumerate(grupos_ordenados)}

                for raw_idx in request.POST.getlist("cons_idx"):
                    idx = str(raw_idx or "").strip()
                    if not idx:
                        continue
                    odf = (request.POST.get(f"cons_{idx}_odf") or "").strip()
                    observacao = (request.POST.get(f"cons_{idx}_observacao") or "").strip()
                    ok = request.POST.get(f"cons_{idx}_ok") == "on"
                    parcial_total = Decimal(str(_parse_decimal(request.POST.get(f"cons_{idx}_parcial") or "0")))
                    data_raw = (request.POST.get(f"cons_{idx}_data") or "").strip()
                    data_val = _parse_date_yyyy_mm_dd(data_raw) if data_raw else None
                    if parcial_total < 0:
                        parcial_total = Decimal("0")

                    matched = grupos_por_idx.get(idx, [])
                    if not matched:
                        continue

                    # O parcial consolidado é distribuído entre os pedidos do produto.
                    restante = parcial_total
                    for item in sorted(matched, key=lambda x: ((x.pedido_numero or ""), x.id)):
                        item.ok = ok
                        item.odf = odf
                        item.observacao = observacao
                        item.data = data_val
                        qtd_item = Decimal(item.quantidade or 0)
                        if restante <= 0:
                            item.parcial = Decimal("0")
                        else:
                            atrib = qtd_item if qtd_item <= restante else restante
                            item.parcial = atrib
                            restante -= atrib
                        item.save()
            else:
                for item in itens_base:
                    prefix = f"item_{item.id}_"
                    item.descricao = (request.POST.get(prefix + "descricao") or "").strip()
                    item.codigo_produto = (request.POST.get(prefix + "codigo_produto") or "").strip()
                    item.odf = (request.POST.get(prefix + "odf") or "").strip()
                    item.observacao = (request.POST.get(prefix + "observacao") or "").strip()
                    item.ok = request.POST.get(prefix + "ok") == "on"
                    parcial_raw = request.POST.get(prefix + "parcial")
                    if parcial_raw is not None and str(parcial_raw).strip() != "":
                        item.parcial = _parse_decimal(parcial_raw)

                    data_raw = request.POST.get(prefix + "data")
                    if data_raw is not None and str(data_raw).strip() != "":
                        item.data = _parse_date_yyyy_mm_dd(data_raw) or item.data

                    quantidade_raw = request.POST.get(prefix + "quantidade")
                    if quantidade_raw is not None and str(quantidade_raw).strip() != "":
                        item.quantidade = _parse_decimal(quantidade_raw)

                    item.save()

            previsao_raw = (request.POST.get("previsao_data") or "").strip()
            liberacao_raw = (request.POST.get("liberacao_producao_data") or "").strip()
            relacao.previsao_data = _parse_date_yyyy_mm_dd(previsao_raw) if previsao_raw else None
            relacao.liberacao_producao_data = _parse_date_yyyy_mm_dd(liberacao_raw) if liberacao_raw else None
            relacao.finalizada = request.POST.get("finalizada") == "on"
            relacao.save()
            return redirect(reverse("relacao_detalhe", kwargs={"relacao_id": relacao.id}))

    ok_filtro = (request.GET.get("ok") or "tudo").strip().lower()
    q = (request.GET.get("q") or "").strip()
    pedido_filtro = (request.GET.get("pedido") or "").strip()
    de = (request.GET.get("de") or "").strip()
    ate = (request.GET.get("ate") or "").strip()

    itens = itens_base
    if ok_filtro == "ok":
        itens = itens.filter(ok=True)
    elif ok_filtro == "pendente":
        itens = itens.filter(ok=False)
    if q:
        itens = itens.filter(
            models.Q(descricao__icontains=q)
            | models.Q(codigo_produto__icontains=q)
            | models.Q(odf__icontains=q)
            | models.Q(observacao__icontains=q)
            | models.Q(pedido_numero__icontains=q)
        )
    if pedido_filtro:
        itens = itens.filter(pedido_numero=pedido_filtro)
    if de:
        data_de = _parse_date_yyyy_mm_dd(de)
        if data_de:
            itens = itens.filter(data__gte=data_de)
    if ate:
        data_ate = _parse_date_yyyy_mm_dd(ate)
        if data_ate:
            itens = itens.filter(data__lte=data_ate)

    m3_rel_filtrado = bool(
        pedido_filtro
        or ok_filtro in ("ok", "pendente")
        or q
        or de
        or ate
    )
    _m3_total_bruto, m3_rel_linhas = _relacao_m3_linhas(itens)
    m3_rel_total = _m3_total_bruto.quantize(Decimal("0.01"))

    visao_consolidada = not bool(pedido_filtro)
    if visao_consolidada:
        agregados: dict[tuple[str, str], dict] = {}
        for it in itens:
            cod = (it.codigo_produto or "").strip()
            desc = (it.descricao or "").strip()
            chave = (cod, desc)
            if chave not in agregados:
                agregados[chave] = {
                    "indice": len(agregados) + 1,
                    "descricao": desc,
                    "codigo_produto": cod,
                    "ok": True,
                    "parcial": Decimal("0"),
                    "quantidade": Decimal("0"),
                    "odf": "",
                    "pedido_numero": "",
                    "data": None,
                    "observacao": "",
                    "id": None,
                }
            agregados[chave]["quantidade"] += Decimal(it.quantidade or 0)
            agregados[chave]["parcial"] += Decimal(it.parcial or 0)
            agregados[chave]["ok"] = bool(agregados[chave]["ok"] and it.ok)
            if not agregados[chave]["odf"] and (it.odf or "").strip():
                agregados[chave]["odf"] = (it.odf or "").strip()
            if not agregados[chave]["observacao"] and (it.observacao or "").strip():
                agregados[chave]["observacao"] = (it.observacao or "").strip()
        itens_exibicao = sorted(
            list(agregados.values()),
            key=lambda x: ((x.get("descricao") or "").lower(), (x.get("codigo_produto") or "").lower()),
        )
    else:
        itens_exibicao = itens

    aba = "finalizadas" if relacao.finalizada else "ativas"
    sidebar_relacoes = Relacao.objects.filter(finalizada=relacao.finalizada).prefetch_related("itens").order_by("-criada_em")[:200]

    contexto = {
        "relacao": relacao,
        "itens": itens_exibicao,
        "visao_consolidada": visao_consolidada,
        "m3_rel_total": m3_rel_total,
        "m3_rel_linhas": m3_rel_linhas,
        "m3_litros_divisor": _litros_divisor_efetivo(),
        "m3_rel_filtrado": m3_rel_filtrado,
        "aba": aba,
        "ok_filtro": ok_filtro,
        "pedido_filtro": pedido_filtro,
        "q": q,
        "de": de,
        "ate": ate,
        "relacao_ativa_id": relacao.id,
        "relacoes": sidebar_relacoes,
    }
    return render(request, "producao/relacao_detalhe.html", contexto)


def relacao_item_toggle_ok(request, relacao_id: int, item_id: int):
    relacao = get_object_or_404(Relacao, pk=relacao_id)
    item = get_object_or_404(RelacaoItem, pk=item_id, relacao=relacao)
    if request.method == "POST":
        item.ok = not item.ok
        item.save()
    return redirect(reverse("relacao_detalhe", kwargs={"relacao_id": relacao.id}))


def relacao_item_delete(request, relacao_id: int, item_id: int):
    relacao = get_object_or_404(Relacao, pk=relacao_id)
    item = get_object_or_404(RelacaoItem, pk=item_id, relacao=relacao)
    if request.method == "POST":
        item.delete()
    return redirect(reverse("relacao_detalhe", kwargs={"relacao_id": relacao.id}))


def relacao_renomear(request, relacao_id: int):
    relacao = get_object_or_404(Relacao, pk=relacao_id)
    if request.method == "POST":
        nome = (request.POST.get("nome") or "").strip()
        if nome:
            relacao.nome = nome
            relacao.save()
    return redirect(reverse("relacoes"))


def relacao_excluir(request, relacao_id: int):
    relacao = get_object_or_404(Relacao, pk=relacao_id)
    if request.method == "POST":
        relacao.delete()
    return redirect(reverse("relacoes"))


def relacoes_visao_geral(request):
    """
    Painel com relações que ainda não atingiram 100% de OK (não finalizadas).
    """
    relacoes_qs = (
        Relacao.objects.filter(finalizada=False)
        .annotate(
            total_itens=Count("itens", distinct=True),
            ok_itens=Count("itens", filter=Q(itens__ok=True), distinct=True),
            total_qtd=Coalesce(
                Sum("itens__quantidade"),
                Value(0, output_field=models.DecimalField(max_digits=18, decimal_places=2)),
            ),
            ok_qtd=Coalesce(
                Sum("itens__quantidade", filter=Q(itens__ok=True)),
                Value(0, output_field=models.DecimalField(max_digits=18, decimal_places=2)),
            ),
        )
        .order_by("-criada_em")[:400]
    )

    cards = []
    for r in relacoes_qs:
        total = int(getattr(r, "total_itens", 0) or 0)
        ok = int(getattr(r, "ok_itens", 0) or 0)
        pct_itens = (ok / total) * 100.0 if total > 0 else 0.0

        total_qtd = float(getattr(r, "total_qtd", 0) or 0)
        ok_qtd = float(getattr(r, "ok_qtd", 0) or 0)
        pct_qtd = (ok_qtd / total_qtd) * 100.0 if total_qtd > 0 else 0.0

        if pct_itens >= 100.0 and pct_qtd >= 100.0:
            continue
        cards.append(
            {
                "id": r.id,
                "nome": r.nome,
                "criada_em": r.criada_em,
                "previsao_data": r.previsao_data,
                "liberacao_producao_data": r.liberacao_producao_data,
                "pedidos_subtitulo": r.pedidos_subtitulo,
                "total_itens": total,
                "ok_itens": ok,
                "pct_itens": pct_itens,
                "pct_qtd": pct_qtd,
                "total_qtd": total_qtd,
                "ok_qtd": ok_qtd,
            }
        )

    n = len(cards)
    compare_rows: list[dict] = []
    for c in cards[:16]:
        nome = c["nome"]
        if len(nome) > 36:
            nome = nome[:35] + "…"
        compare_rows.append(
            {
                "nome": nome,
                "pct_qtd": round(float(c["pct_qtd"]), 1),
                "pct_itens": round(float(c["pct_itens"]), 1),
            }
        )
    vg_media_pct_qtd = vg_media_pct_itens = None
    if n:
        vg_media_pct_qtd = round(sum(float(c["pct_qtd"]) for c in cards) / n, 1)
        vg_media_pct_itens = round(sum(float(c["pct_itens"]) for c in cards) / n, 1)

    chart_h = max(200, min(520, 48 + len(compare_rows) * 34)) if compare_rows else 0

    contexto = {
        "cards": cards,
        "compare_chart_json": json.dumps(compare_rows),
        "vg_count": n,
        "vg_media_pct_qtd": vg_media_pct_qtd,
        "vg_media_pct_itens": vg_media_pct_itens,
        "compare_chart_height": chart_h,
    }
    return render(request, "producao/relacoes_visao_geral.html", contexto)


def _relacoes_compras_list() -> list[dict]:
    """Relações + itens (até 200) para Compras / Necessidades; inclui «ok» por item."""
    relacoes_qs = (
        Relacao.objects.all()
        .order_by("-criada_em")
        .prefetch_related("itens")
    )[:200]

    relacoes_payload: list[dict] = []
    for r in relacoes_qs:
        itens_payload: list[dict] = []
        for it in r.itens.all().order_by("descricao", "codigo_produto", "id"):
            cod_prod = (it.codigo_produto or "").strip()
            if not cod_prod:
                continue
            itens_payload.append(
                {
                    "codigo_produto": cod_prod,
                    "descricao_produto": (it.descricao or "").strip(),
                    "quantidade": float(it.quantidade or 0),
                    "ok": bool(it.ok),
                }
            )

        relacoes_payload.append(
            {
                "id": r.id,
                "nome": r.nome,
                "finalizada": bool(r.finalizada),
                "itens": itens_payload,
            }
        )

    return relacoes_payload


def _relacoes_json_payload() -> str:
    """JSON das relações + itens (até 200 mais recentes) para cálculo de necessidades no navegador."""
    return json.dumps(_relacoes_compras_list())


def _normalize_code_key_estoque(raw: str) -> str:
    raw = (raw or "").strip()
    if not raw:
        return ""
    only_digits = re.sub(r"\D", "", raw)
    if only_digits:
        try:
            return str(int(only_digits))
        except ValueError:
            return only_digits
    return re.sub(r"\s+", "", raw.upper())


def _to_number_estoque(val) -> float:
    if val is None:
        return 0.0
    s = str(val).strip()
    if not s:
        return 0.0
    has_comma = "," in s
    has_dot = "." in s
    normalized = re.sub(r"\s", "", s)
    if has_comma and has_dot:
        normalized = normalized.replace(".", "").replace(",", ".")
    elif has_comma:
        normalized = normalized.replace(",", ".")
    try:
        return float(normalized)
    except ValueError:
        return 0.0


def _parse_estoque_rows_matrix(rows: list) -> dict[str, float]:
    estoque_map: dict[str, float] = defaultdict(float)
    current_code = ""
    current_saldo: float | None = None
    for row in rows:
        cells = list(row) if row else []
        line = " ".join(str(c if c is not None else "") for c in cells)
        m_prod = re.search(r"Produto:\s*([0-9A-Za-z]+)", line, re.I)
        if m_prod:
            current_code = _normalize_code_key_estoque(m_prod.group(1))
            current_saldo = None
        m_saldo = re.search(r"Saldo:\s*([0-9.,-]+)", line, re.I)
        if m_saldo and current_code:
            current_saldo = _to_number_estoque(m_saldo.group(1))
            estoque_map[current_code] += current_saldo
        if re.search(r"Sub-Total", line, re.I) and current_code and (
            current_saldo is None or current_saldo == 0
        ):
            nums = [_to_number_estoque(v) for v in cells]
            nums = [v for v in nums if v > 0]
            if nums:
                estoque_map[current_code] += nums[-1]
    return dict(estoque_map)


def _estoque_saved_dir() -> Path:
    root = getattr(settings, "MEDIA_ROOT", None)
    if not root:
        root = Path(settings.BASE_DIR) / "media"
    return Path(root) / "carcacas_estoque"


def _estoque_saved_json_path() -> Path:
    return _estoque_saved_dir() / "estoque_latest.json"


def _estoque_meta_original_name_path() -> Path:
    return _estoque_saved_dir() / "estoque_latest_original.txt"


def _parse_estoque_excel_bytes(data: bytes, filename: str) -> tuple[dict[str, float], str | None]:
    del filename  # reservado para mensagens futuras
    if not data:
        return {}, "Arquivo vazio."
    rows: list[list] = []
    try:
        if len(data) >= 2 and data[:2] == b"PK":
            from io import BytesIO

            from openpyxl import load_workbook

            wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    rows.append(list(row))
            wb.close()
        else:
            import xlrd

            wb = xlrd.open_workbook(file_contents=data)
            for sheet_idx in range(wb.nsheets):
                sh = wb.sheet_by_index(sheet_idx)
                for r in range(sh.nrows):
                    rows.append([sh.cell_value(r, c) for c in range(sh.ncols)])
    except Exception as exc:  # noqa: BLE001
        return {}, f"Não foi possível ler o arquivo: {exc}"
    est_map = _parse_estoque_rows_matrix(rows)
    if not est_map:
        return {}, "Nenhum saldo foi identificado no arquivo (verifique o formato de exportação)."
    return est_map, None


def _load_estoque_map_from_disk() -> dict[str, float]:
    p = _estoque_saved_json_path()
    if not p.is_file():
        return {}
    try:
        raw = json.loads(p.read_text(encoding="utf-8"))
        if not isinstance(raw, dict):
            return {}
        out: dict[str, float] = {}
        for k, v in raw.items():
            try:
                out[str(k)] = float(v)
            except (TypeError, ValueError):
                continue
        return out
    except (OSError, json.JSONDecodeError):
        return {}


def _save_estoque_map_to_disk(est_map: dict[str, float], original_filename: str) -> str | None:
    try:
        d = _estoque_saved_dir()
        d.mkdir(parents=True, exist_ok=True)
        dest = _estoque_saved_json_path()
        tmp = dest.with_suffix(".tmp")
        tmp.write_text(json.dumps(est_map, ensure_ascii=False, sort_keys=True), encoding="utf-8")
        tmp.replace(dest)
        try:
            _estoque_meta_original_name_path().write_text(
                (original_filename or "estoque.xlsx").strip() or "estoque.xlsx",
                encoding="utf-8",
            )
        except OSError:
            pass
        return None
    except OSError as exc:
        return str(exc)


@ensure_csrf_cookie
def necessidades_carcacas(request):
    """
    Gera a tela de "Necessidades de Carcaças" a partir de:
    - relações cadastradas no módulo "Relações" (Relacao + RelacaoItem)
    - uma base de dados importada pelo usuário no frontend (Excel -> localStorage)

    A comparação é feita por:
    - Cod_Interno (-> RelacaoItem.codigo_produto)
    """

    contexto = {
        "relacoes_json": _relacoes_json_payload(),
        "estoque_servidor_json": json.dumps(_load_estoque_map_from_disk()),
    }
    return render(request, "producao/necessidades_carcacas.html", contexto)


@never_cache
@ensure_csrf_cookie
def comprar(request):
    """
    Painel de apoio à compra de acessórios: cruzar estoque (importado) com consumo
    agregado das relações.
    """
    return render(
        request,
        "producao/comprar.html",
        {
            "comprar_ui_rev": "2026-04-06-compras-ok-neg-alerta",
            "relacoes_json": _relacoes_json_payload(),
            "estoque_servidor_json": json.dumps(_load_estoque_map_from_disk()),
        },
    )


@never_cache
def api_relacoes_compras(request):
    return JsonResponse(_relacoes_compras_list(), safe=False)


@never_cache
def api_estoque_carcacas_json(request):
    return JsonResponse(_load_estoque_map_from_disk())


DEFAULT_MRP_XLS_PATH = os.environ.get(
    "MRP_XLS_PATH",
    r"h:\Meu Drive\PCP\EXP_MRP\MRP.xls",
)


def _mrp_saved_abs_path() -> Path:
    root = getattr(settings, "MEDIA_ROOT", None)
    if not root:
        root = Path(settings.BASE_DIR) / "media"
    return Path(root) / "mrp" / "MRP_latest.xls"


def _mrp_meta_original_name_path() -> Path:
    return _mrp_saved_abs_path().parent / "MRP_latest_original.txt"


def _save_mrp_upload(data: bytes, original_filename: str) -> tuple[Path | None, str | None]:
    dest = _mrp_saved_abs_path()
    try:
        dest.parent.mkdir(parents=True, exist_ok=True)
        tmp = dest.with_suffix(".tmp")
        tmp.write_bytes(data)
        tmp.replace(dest)
        try:
            _mrp_meta_original_name_path().write_text(
                (original_filename or "MRP.xls").strip() or "MRP.xls",
                encoding="utf-8",
            )
        except OSError:
            pass
        return dest, None
    except OSError as exc:
        return None, str(exc)


def _read_mrp_saved_display_name() -> str:
    p = _mrp_meta_original_name_path()
    if p.is_file():
        return (p.read_text(encoding="utf-8") or "").strip()
    return ""


def _normalize_odf_key(val) -> str:
    """Normaliza ODF vindos do Excel (número) ou da relação (texto) para comparação."""
    if val is None or val == "":
        return ""
    if isinstance(val, float):
        if val != val:  # NaN
            return ""
        if val == int(val):
            return str(int(val))
        s = ("%s" % val).strip()
        return s.rstrip("0").rstrip(".") if "." in s else s
    s = str(val).strip().replace(" ", "")
    if re.fullmatch(r"\d+\.0+", s):
        return s.split(".")[0]
    if re.fullmatch(r"\d+$", s):
        try:
            return str(int(s))
        except ValueError:
            return s
    if re.fullmatch(r"\d+[.,]\d+$", s):
        try:
            f = float(s.replace(",", "."))
            if f == int(f):
                return str(int(f))
        except (ValueError, OverflowError):
            pass
    return s


def _expand_odf_keys_for_match(keys: set[str]) -> set[str]:
    """Inclui variantes numéricas (só dígitos, int) para cruzar MRP × relação."""
    out: set[str] = set()
    for k in keys:
        if not k:
            continue
        out.add(k)
        digits = re.sub(r"\D", "", k)
        if digits:
            out.add(digits)
            try:
                out.add(str(int(digits)))
            except ValueError:
                pass
    return {x for x in out if x}


def _relation_odf_match_keys(o_raw: str) -> set[str]:
    n = _normalize_odf_key(o_raw)
    out = {n} if n else set()
    digits = re.sub(r"\D", "", (o_raw or "").strip())
    if digits:
        out.add(digits)
        try:
            out.add(str(int(digits)))
        except ValueError:
            pass
    return {x for x in out if x}


def _detect_mrp_odf_col_index(sh) -> int:
    """Usa cabeçalho com 'ODF' na 1ª linha; senão coluna G (índice 6)."""
    if sh.nrows > 0 and sh.ncols > 0:
        for j in range(sh.ncols):
            cell = sh.cell_value(0, j)
            h = str(cell).strip().lower()
            if "odf" in h:
                return j
    return 6


def _load_mrp_odf_set_from_xls(*, path: str | None = None, file_contents: bytes | None = None) -> tuple[set[str], str | None]:
    try:
        import xlrd
    except ImportError:
        return set(), "Instale xlrd: pip install xlrd"
    try:
        if file_contents is not None:
            wb = xlrd.open_workbook(file_contents=file_contents)
        elif path:
            wb = xlrd.open_workbook(path)
        else:
            return set(), "Nenhum arquivo MRP informado."
    except Exception as exc:  # noqa: BLE001
        return set(), f"Não foi possível ler o MRP: {exc}"

    sh = wb.sheet_by_index(0)
    col_odf = _detect_mrp_odf_col_index(sh)
    if col_odf >= sh.ncols:
        col_odf = max(0, sh.ncols - 1)
    odfs: set[str] = set()
    start_row = 1  # linha 0 = cabeçalho; se só existir cabeçalho, não há linhas de dados
    for r in range(start_row, sh.nrows):
        if col_odf >= sh.ncols:
            break
        k = _normalize_odf_key(sh.cell_value(r, col_odf))
        if k:
            odfs.add(k)
    return odfs, None


def _process_mrp_post_upload(request) -> tuple[bool, str | None]:
    upload = request.FILES.get("mrp")
    if not upload:
        return False, "Selecione um arquivo MRP (.xls) para enviar."
    data = upload.read()
    _, err = _load_mrp_odf_set_from_xls(file_contents=data)
    if err:
        return False, err
    _, save_err = _save_mrp_upload(data, upload.name)
    if save_err:
        return False, f"MRP carregado na memória, mas não foi possível salvar no servidor: {save_err}"
    return True, None


@require_POST
def pcp_import_mrp(request):
    ok, err = _process_mrp_post_upload(request)
    if ok:
        return JsonResponse({"ok": True, "message": "MRP salvo no servidor."})
    return JsonResponse({"ok": False, "error": err or "Erro ao importar MRP."}, status=400)


@require_POST
def pcp_import_estoque(request):
    upload = request.FILES.get("estoque")
    if not upload:
        return JsonResponse({"ok": False, "error": "Nenhum arquivo enviado."}, status=400)
    data = upload.read()
    est_map, parse_err = _parse_estoque_excel_bytes(data, upload.name)
    if parse_err:
        return JsonResponse({"ok": False, "error": parse_err}, status=400)
    save_err = _save_estoque_map_to_disk(est_map, upload.name)
    if save_err:
        return JsonResponse({"ok": False, "error": save_err}, status=500)
    return JsonResponse(
        {
            "ok": True,
            "message": f"Estoque salvo no servidor ({len(est_map)} código(s)).",
            "count": len(est_map),
        }
    )


def odfs_movimentar(request):
    """
    Lista itens com OK nas relações cujo ODF aparece no MRP (.xls).
    Coluna: cabeçalho "ODF" na 1ª linha, ou coluna G (índice 6).
    Quantidade a movimentar: Parcial (se > 0), senão Total.
    """
    erro = None
    mrp_source = ""
    mrp_odfs: set[str] = set()
    rows: list[dict] = []
    rows_for_needs: list[dict] = []
    odf_diag: str | None = None
    sample_rel_odfs: list[str] = []
    sample_mrp_odfs: list[str] = []

    if request.method == "POST":
        ok, err_msg = _process_mrp_post_upload(request)
        if ok:
            return redirect(f"{reverse('odfs_movimentar')}?saved=1")
        erro = err_msg
    elif request.method == "GET":
        saved = _mrp_saved_abs_path()
        if saved.is_file():
            mrp_odfs, err = _load_mrp_odf_set_from_xls(path=str(saved))
            if err:
                erro = err
            else:
                orig = _read_mrp_saved_display_name()
                mrp_source = (
                    f"Último upload salvo ({saved.name}"
                    + (f", enviado como «{orig}»" if orig else "")
                    + ")"
                )
        elif os.path.isfile(DEFAULT_MRP_XLS_PATH):
            mrp_odfs, err = _load_mrp_odf_set_from_xls(path=DEFAULT_MRP_XLS_PATH)
            if err:
                erro = err
            else:
                mrp_source = DEFAULT_MRP_XLS_PATH
        else:
            erro = None

    mrp_expanded = _expand_odf_keys_for_match(mrp_odfs) if mrp_odfs else set()

    if not erro and mrp_expanded:
        itens = (
            RelacaoItem.objects.filter(ok=True)
            .select_related("relacao")
            .order_by("relacao__nome", "indice", "id")
        )
        rel_odfs_seen: list[str] = []
        for it in itens:
            o_raw = (it.odf or "").strip()
            if not o_raw:
                continue
            keys = _relation_odf_match_keys(o_raw)
            if len(rel_odfs_seen) < 8 and o_raw not in rel_odfs_seen:
                rel_odfs_seen.append(o_raw)
            if not keys.intersection(mrp_expanded):
                continue
            parcial = float(it.parcial or 0)
            qtd = float(it.quantidade or 0)
            q_mov = parcial if parcial > 0 else qtd
            rows.append(
                {
                    "relacao_nome": it.relacao.nome,
                    "produto_nome": (it.descricao or "").strip() or (it.codigo_produto or "").strip() or "—",
                    "codigo_produto": (it.codigo_produto or "").strip(),
                    "odf": o_raw,
                    "quantidade_mov": q_mov,
                    "is_parcial": parcial > 0,
                    "quantidade_total": qtd,
                }
            )
            # payload resumido para análise de carcaças/discos a partir das ODFs a movimentar
            if (it.codigo_produto or "").strip() and q_mov > 0:
                rows_for_needs.append(
                    {
                        "codigo_produto": (it.codigo_produto or "").strip(),
                        "quantidade_mov": q_mov,
                    }
                )

        if not rows and mrp_odfs:
            sample_rel_odfs = rel_odfs_seen[:5]
            sample_mrp_odfs = sorted(mrp_odfs, key=lambda x: (len(str(x)), str(x)))[:5]
            ok_com_odf = RelacaoItem.objects.filter(ok=True).exclude(odf__isnull=True).exclude(odf="").count()
            if ok_com_odf == 0:
                odf_diag = "Não há itens com OK e ODF preenchido nas relações."
            elif sample_rel_odfs:
                odf_diag = (
                    f"Há {ok_com_odf} item(ns) com OK e ODF, mas nenhum bate com o MRP "
                    "(confira zeros à esquerda, formato ou se o arquivo é o export correto)."
                )

    upload_saved_notice = (request.GET.get("saved") or "").strip() == "1"

    contexto = {
        "erro": erro,
        "rows": rows,
        "mrp_source": mrp_source,
        "mrp_odf_count": len(mrp_odfs),
        "rows_count": len(rows),
        "default_mrp_path": DEFAULT_MRP_XLS_PATH,
        "mrp_saved_path": str(_mrp_saved_abs_path()),
        "odf_diag": odf_diag,
        "sample_rel_odfs": sample_rel_odfs,
        "sample_mrp_odfs": sample_mrp_odfs,
        "upload_saved_notice": upload_saved_notice,
        # usado em odfs_movimentar.html para cruzar com base/estoque de carcaças
        "rows_for_needs_json": json.dumps(rows_for_needs),
    }
    return render(request, "producao/odfs_movimentar.html", contexto)


def _parse_date_yyyy_mm_dd(raw: str):
    try:
        return datetime.strptime(raw, "%Y-%m-%d").date()
    except Exception:
        return None


def _looks_like_odf_token(s: str, *, min_digits: int = 4) -> bool:
    """
    Heurística: ODF costuma ser inteiro longo; quantidade costuma ter decimal (10,5) ou ser curta.
    Evita confundir Parcial|Total (2 colunas) com Total|ODF.
    """
    txt = (s or "").strip().replace(" ", "")
    if not txt:
        return False
    if _parse_date_br_or_iso(txt):
        return False
    # Quantidade decimal típica
    if re.fullmatch(r"\d+[.,]\d+$", txt):
        return False
    if re.fullmatch(r"\d{1,3}(?:\.\d{3})+(?:,\d+)?$", txt):
        return False
    digits = re.sub(r"\D", "", txt)
    if not digits.isdigit():
        return False
    return len(digits) >= min_digits


def _split_tail_parcial_total_odf(tail: list[str]) -> tuple[str, str, str]:
    """Interpreta o trecho após OK: Parcial | Total | ODF (ODF pode faltar)."""
    if not tail:
        return "0", "0", ""
    if len(tail) == 1:
        t0 = tail[0]
        # Só trata como ODF “sozinho” se for número longo (evita confundir com total 1000).
        if _looks_like_odf_token(t0, min_digits=5):
            return "0", "0", t0
        return "0", t0, ""
    if len(tail) == 2:
        a, b = tail[0], tail[1]
        if _looks_like_odf_token(b, min_digits=4) and not _looks_like_odf_token(a, min_digits=4):
            return "0", a, b
        return a, b, ""
    return tail[-3], tail[-2], tail[-1]


def _pick_odf_from_unused_columns(cols: list[str], colmap: dict) -> str:
    """Se o cabeçalho não nomeou ODF, tenta a última coluna numérica longa não usada."""
    used: set[int] = set()
    for key in ("descricao", "codigo", "ok", "parcial", "total", "data"):
        i = colmap.get(key)
        if i is not None and isinstance(i, int) and i >= 0:
            used.add(i)
    for idx in range(len(cols) - 1, -1, -1):
        if idx in used:
            continue
        v = (cols[idx] or "").strip()
        if not v or _parse_date_br_or_iso(v):
            continue
        if _looks_like_odf_token(v, min_digits=5):
            return v
    return ""


def _parse_colagem_excel(texto: str):
    """
    Aceita colagens comuns do Excel.

    Formato atual (pedido):
    Descrição | Cód. do Produto | OK | Parcial | Total | ODF

    Formato antigo (ainda aceito):
    Descrição | Cód. do Produto | OK | Total | ODF | Data
    """
    linhas = [l for l in (texto or "").splitlines() if l.strip()]
    if not linhas:
        return []

    # Excel geralmente cola como TSV (tab-separated). Se vier com ';' também aceitamos.
    def split_cols(linha: str):
        if "\t" in linha:
            return [c.strip() for c in linha.split("\t")]
        if ";" in linha:
            return [c.strip() for c in linha.split(";")]
        # fallback: algumas colagens vêm com colunas separadas por múltiplos espaços
        return [c.strip() for c in re.split(r"\s{2,}", linha.strip())]

    def parse_sem_delimitador(linha: str):
        """
        Fallback para colagens sem TAB/; e sem espaços duplos (apenas espaços comuns).
        Baseado no padrão visto (imagem):
        Descrição ... | CÓDIGO_NUMÉRICO | OK | (PARCIAL) | TOTAL | ODF
        Obs: ODF pode ser numérico, então não dá pra usar "último número = total".
        """
        s = (linha or "").strip()
        if not s:
            return None

        m = re.search(r"(?P<codigo>\d{6,})", s)
        if not m:
            return None

        descricao = s[: m.start()].strip()
        codigo = m.group("codigo")
        resto = s[m.end() :].strip()

        toks = [t for t in re.split(r"\s+", resto) if t]
        ok_raw = toks[0] if toks else ""
        if ok_raw.strip().lower() in {"1", "true", "sim", "ok", "x", "y", "yes"}:
            toks = toks[1:]
        else:
            ok_raw = ""

        # Depois do OK: Parcial | Total | ODF (Parcial pode estar vazio)
        if len(toks) >= 3:
            parcial_raw, total_raw, odf = toks[-3], toks[-2], toks[-1]
        elif len(toks) == 2:
            a, b = toks[-2], toks[-1]
            if _looks_like_odf_token(b, min_digits=4) and not _looks_like_odf_token(a, min_digits=4):
                parcial_raw, total_raw, odf = "0", a, b
            else:
                parcial_raw, total_raw, odf = a, b, ""
        elif len(toks) == 1:
            t = toks[-1]
            if _looks_like_odf_token(t, min_digits=5):
                parcial_raw, total_raw, odf = "0", "0", t
            else:
                parcial_raw, total_raw, odf = "0", t, ""
        else:
            parcial_raw, total_raw, odf = "0", "0", ""

        return {
            "descricao": descricao,
            "codigo_produto": codigo,
            "ok_raw": ok_raw,
            "parcial_raw": parcial_raw,
            "total_raw": total_raw,
            "odf": odf,
            "data_raw": "",
        }

    header = [c.strip() for c in split_cols(linhas[0])]
    header_l = [c.lower() for c in header]
    tem_descr = any("descr" in c for c in header_l)
    tem_cod = any(("cód" in c) or ("cod" in c) for c in header_l)
    tem_ok = any("ok" in c for c in header_l)
    parece_cabecalho = (tem_descr and tem_cod) or (tem_ok and tem_cod and len(header) >= 5)

    colmap = None
    if parece_cabecalho:
        def find_idx(preds: list[str]):
            for i, c in enumerate(header_l):
                for p in preds:
                    if p in c:
                        return i
            return None

        def find_odf_idx():
            for i, c in enumerate(header_l):
                if "odf" in c:
                    return i
            for i, c in enumerate(header_l):
                if "ordem" in c and ("fab" in c or "fabric" in c):
                    return i
            for i, c in enumerate(header_l):
                if "fabricação" in c or "fabricacao" in c:
                    return i
            return None

        def find_total_idx():
            """
            Coluna de quantidade (Total).

            Não usar o predicado 'quant' em find_idx: ele casa com 'quantidade'
            e, se essa coluna vier antes de 'Total', o import lia o valor errado
            (ou vazio) — na colagem do Excel os números continuam visíveis.
            """
            for i, c in enumerate(header_l):
                if re.search(r"\btotal\b", c, re.I):
                    return i
            for i, c in enumerate(header_l):
                if "quantidade" in c:
                    return i
            for i, c in enumerate(header_l):
                if re.search(r"\bqtd\b", c, re.I):
                    return i
            return None

        colmap = {
            "descricao": find_idx(["descr"]),
            "codigo": find_idx(["cód", "cod", "produto"]),
            "ok": find_idx(["ok"]),
            "parcial": find_idx(["parcial"]),
            "total": find_total_idx(),
            "odf": find_odf_idx(),
            "data": find_idx(["data"]),
        }
        linhas = linhas[1:]

    itens = []
    for linha in linhas:
        cols = split_cols(linha)
        # Excel às vezes inclui TAB/coluna vazia no final do range copiado.
        # Remover vazios finais evita deslocar Total/ODF.
        while cols and cols[-1] == "":
            cols.pop()
        if len(cols) < 2:
            parsed = parse_sem_delimitador(linha)
            if not parsed:
                continue

            ok = parsed["ok_raw"].strip().lower() in {"1", "true", "sim", "ok", "x", "y", "yes"}
            parcial = _parse_decimal(parsed["parcial_raw"])
            quantidade = _parse_decimal(parsed["total_raw"])
            data = _parse_date_br_or_iso(parsed["data_raw"])

            itens.append(
                {
                    "descricao": parsed["descricao"],
                    "codigo_produto": parsed["codigo_produto"],
                    "ok": ok,
                    "parcial": parcial,
                    "quantidade": quantidade,
                    "odf": (parsed["odf"] or "").strip(),
                    "data": data,
                }
            )
            continue

        def get_by_idx(i, default=""):
            if i is None:
                return default
            if i < 0 or i >= len(cols):
                return default
            return cols[i]

        if colmap:
            descricao = get_by_idx(colmap["descricao"], "")
            codigo = get_by_idx(colmap["codigo"], "")
            ok_raw = get_by_idx(colmap["ok"], "")
            parcial_raw = get_by_idx(colmap.get("parcial"), "0")
            total_raw = get_by_idx(colmap["total"], "0")
            odf = get_by_idx(colmap["odf"], "")
            if not (odf or "").strip() and colmap.get("odf") is None:
                fb = _pick_odf_from_unused_columns(cols, colmap)
                if fb:
                    odf = fb
            data_raw = get_by_idx(colmap["data"], "")
        else:
            descricao = cols[0] if len(cols) > 0 else ""
            codigo = cols[1] if len(cols) > 1 else ""
            ok_raw = cols[2] if len(cols) > 2 else ""
            # Sem cabeçalho: o Excel costuma trazer várias colunas vazias.
            # Estratégia: após a coluna OK, considerar apenas valores não vazios.
            # Assim conseguimos ler sempre do fim: Parcial | Total | ODF
            tail = [c for c in (cols[3:] if len(cols) > 3 else []) if c != ""]
            data_raw = ""

            # Se a última coluna for data, remove do "tail" antes de calcular
            # Parcial | Total | ODF. Isso evita perder o Total em colagens onde
            # a data vem ao final e a pré-visualização está correta.
            if tail and _parse_date_br_or_iso(tail[-1]):
                data_raw = tail.pop()

            parcial_raw, total_raw, odf = _split_tail_parcial_total_odf(tail)

            # Se vierem colunas a mais, tenta achar uma data em qualquer coluna extra.
            if not data_raw and len(cols) > 6:
                for v in cols[6:]:
                    if _parse_date_br_or_iso(v):
                        data_raw = v
                        break

        ok = ok_raw.strip().lower() in {"1", "true", "sim", "ok", "x", "y", "yes"}

        parcial = _parse_decimal(parcial_raw)
        quantidade = _parse_decimal(total_raw)
        data = _parse_date_br_or_iso(data_raw)

        itens.append(
            {
                "descricao": descricao,
                "codigo_produto": codigo,
                "ok": ok,
                "parcial": parcial,
                "quantidade": quantidade,
                "odf": (odf or "").strip(),
                "data": data,
            }
        )

    if not itens:
        raise ValueError("Nenhuma linha válida encontrada na colagem.")
    return itens


def _parse_decimal(raw: str):
    txt = (raw or "").strip()
    if not txt:
        return 0

    # Remove moeda/unidades e mantém apenas números/separadores comuns.
    allowed = set("0123456789-.,")
    txt = "".join(ch for ch in txt if ch in allowed).strip()
    if not txt:
        return 0

    # Se houver vírgula e ponto, detecta qual é o separador decimal pelo último símbolo.
    if "," in txt and "." in txt:
        if txt.rfind(",") > txt.rfind("."):
            # 1.234,56 -> 1234.56
            txt = txt.replace(".", "").replace(",", ".")
        else:
            # 1,234.56 -> 1234.56
            txt = txt.replace(",", "")
    elif "," in txt:
        # Apenas vírgula: se parecer milhar (1,234), remove; senão trata como decimal.
        parts = txt.split(",")
        if len(parts) > 1 and len(parts[-1]) == 3 and all(part.isdigit() for part in parts):
            txt = "".join(parts)
        else:
            txt = txt.replace(",", ".")
    elif "." in txt:
        # Apenas ponto: se parecer milhar (1.234), remove; senão mantém como decimal.
        parts = txt.split(".")
        if len(parts) > 1 and len(parts[-1]) == 3 and all(part.isdigit() for part in parts):
            txt = "".join(parts)

    try:
        return float(txt)
    except Exception:
        return 0


def _parse_date_br_or_iso(raw: str):
    txt = (raw or "").strip()
    if not txt:
        return None
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(txt, fmt).date()
        except Exception:
            continue
    return None


def distribuir_operacao(request):
    categorias = Categoria.objects.all().order_by("nome")
    relacoes = Relacao.objects.all().order_by("-criada_em")[:200]

    categorias_atuais = []
    relacao_atual = None
    resumo = []

    cats_raw = (request.GET.get("cats") or "").strip()
    cat_id = request.GET.get("cat")
    rel_id = request.GET.get("rel")

    cats_ids: list[int] = []
    if cats_raw:
        for p in cats_raw.split(","):
            try:
                cats_ids.append(int(p.strip()))
            except Exception:
                continue
    elif cat_id:
        try:
            cats_ids = [int(cat_id)]
        except Exception:
            cats_ids = []

    if cats_ids and categorias:
        categorias_atuais = list(categorias.filter(id__in=cats_ids))

    if rel_id and relacoes:
        relacao_atual = Relacao.objects.filter(id=rel_id).first()

    if categorias_atuais and relacao_atual:
        codigos = set(
            Produto.objects.filter(categorias__in=categorias_atuais).values_list("codigo", flat=True).distinct()
        )
        itens = relacao_atual.itens.all()
        itens_filtrados = [i for i in itens if (i.codigo_produto or "").strip() in codigos]

        agreg = {}
        for it in itens_filtrados:
            chave = (it.descricao or "").strip() or (it.codigo_produto or "").strip() or "Sem descrição"
            agreg[chave] = agreg.get(chave, 0) + float(it.quantidade or 0)

        resumo = [{"descricao": k, "total": v} for k, v in sorted(agreg.items(), key=lambda x: x[0].lower())]

    cats_set = set(cats_ids)
    categorias_ui = []
    for c in list(categorias):
        # toggle
        next_ids = [x for x in cats_ids if x != c.id] if c.id in cats_set else (cats_ids + [c.id])
        next_ids = sorted(set([x for x in next_ids if x]))
        qs = []
        if next_ids:
            qs.append("cats=" + ",".join(str(x) for x in next_ids))
        if relacao_atual:
            qs.append("rel=" + str(relacao_atual.id))
        href = ("?" + "&".join(qs)) if qs else "?"
        categorias_ui.append(
            {
                "id": c.id,
                "nome": c.nome,
                "active": c.id in cats_set,
                "href": href,
            }
        )

    contexto = {
        "categorias": categorias,
        "categorias_ui": categorias_ui,
        "relacoes": relacoes,
        "categorias_atuais": categorias_atuais,
        "cats_ids": cats_ids,
        "relacao_atual": relacao_atual,
        "resumo": resumo,
    }
    return render(request, "producao/distribuir.html", contexto)


def distribuir_imprimir(request):
    categorias = Categoria.objects.all()
    cats_raw = (request.GET.get("cats") or "").strip()
    cat_id = request.GET.get("cat")
    rel_id = request.GET.get("rel")

    cats_ids: list[int] = []
    if cats_raw:
        for p in cats_raw.split(","):
            try:
                cats_ids.append(int(p.strip()))
            except Exception:
                continue
    elif cat_id:
        try:
            cats_ids = [int(cat_id)]
        except Exception:
            cats_ids = []

    categorias_atuais = list(categorias.filter(id__in=cats_ids)) if cats_ids else []
    relacao_atual = Relacao.objects.filter(id=rel_id).first() if rel_id else None
    entregue_por = (request.GET.get("entregue_por") or "").strip()

    blocos = []
    if categorias_atuais and relacao_atual:
        itens = list(relacao_atual.itens.all())
        for cat in categorias_atuais:
            codigos = set(cat.produtos.values_list("codigo", flat=True))
            itens_filtrados = [i for i in itens if (i.codigo_produto or "").strip() in codigos]
            agreg = {}
            for it in itens_filtrados:
                chave = (it.descricao or "").strip() or (it.codigo_produto or "").strip() or "Sem descrição"
                agreg[chave] = agreg.get(chave, 0) + float(it.quantidade or 0)
            resumo = [{"descricao": k, "total": v} for k, v in sorted(agreg.items(), key=lambda x: x[0].lower())]
            blocos.append({"categoria": cat, "resumo": resumo})

    contexto = {
        "categorias": categorias_atuais,
        "relacao": relacao_atual,
        "blocos": blocos,
        "entregue_por": entregue_por,
    }
    return render(request, "producao/distribuir_imprimir.html", contexto)


def configurar_distribuicao(request):
    if request.method == "POST":
        acao = request.POST.get("acao")
        if acao == "add_categoria":
            nome = (request.POST.get("nome") or "").strip()
            if nome:
                Categoria.objects.get_or_create(nome=nome)
            return redirect(reverse("configurar_distribuicao"))
        if acao == "set_produto_categorias":
            produto_id = request.POST.get("produto_id")
            if produto_id:
                produto = get_object_or_404(Produto, pk=produto_id)
                cat_ids = request.POST.getlist("categorias")
                try:
                    cat_ids = [int(x) for x in cat_ids]
                except Exception:
                    cat_ids = []
                produto.categorias.set(Categoria.objects.filter(id__in=cat_ids))
            return redirect(reverse("configurar_distribuicao"))

    categorias = Categoria.objects.all().order_by("nome")
    produtos = Produto.objects.prefetch_related("categorias").order_by("nome")
    produtos_sem_categoria = produtos.filter(categorias__isnull=True)
    categorias_com_contagem = []
    for c in categorias:
        categorias_com_contagem.append(
            type(
                "CatWithCount",
                (),
                {"id": c.id, "nome": c.nome, "qtd_produtos": c.produtos.count()},
            )
        )

    contexto = {
        "categorias": categorias_com_contagem,
        "categorias_raw": list(categorias),
        "produtos": produtos,
        "produtos_sem_categoria": produtos_sem_categoria,
    }
    return render(request, "producao/configurar_distribuicao.html", contexto)


@never_cache
def pintura_epox(request):
    # pintura_ui_rev: altere ao mudar o template para confirmar no browser que não há cache antigo
    rel_links = list(Relacao.objects.order_by("-criada_em").values("id", "nome")[:500])
    return render(
        request,
        "producao/pintura_epox.html",
        {
            "pintura_ui_rev": "2026-04-06-r1-rel-links-compras-fs-filtro",
            "relacoes_link_json": json.dumps(rel_links, ensure_ascii=False),
        },
    )


def controle_producao_login(request):
    labels = ["1° Turno", "2° Turno", "3° Turno"]
    valores = [0, 0, 0]

    # Não existe modelo de apontamento por turno neste projeto.
    # Então o gráfico inicia vazio e só mostra dados reais quando houver
    # integração com uma fonte de produção por turno.
    contexto = {
        "turno_labels_json": json.dumps(labels, ensure_ascii=False),
        "turno_valores_json": json.dumps(valores),
        "tem_dados_turno": any(v > 0 for v in valores),
    }
    return render(request, "producao/controle_producao_login.html", contexto)


def _dec_or_zero(value) -> Decimal:
    s = str(value or "").strip().replace(",", ".")
    if not s:
        return Decimal("0")
    try:
        return Decimal(s)
    except Exception:  # noqa: BLE001
        return Decimal("0")


def _time_or_none(value):
    s = str(value or "").strip()
    if not s:
        return None
    try:
        return datetime.strptime(s, "%H:%M").time()
    except ValueError:
        return None


def _serialize_cacarola_registro(r: CacarolaRegistro) -> dict:
    return {
        "uid": r.uid,
        "data": r.data.isoformat() if r.data else "",
        "turno": r.turno or "",
        "maquina": r.maquina or "",
        "responsavel": r.responsavel or "",
        "tipoProduto": r.tipo_produto or "",
        "produto": r.produto or "",
        "quantidade": float(r.quantidade or 0),
        "odf": r.odf or "",
        "inicio": r.inicio.strftime("%H:%M") if r.inicio else "",
        "fim": r.fim.strftime("%H:%M") if r.fim else "",
        "refeicao": float(r.refeicao or 0),
        "tempo": float(r.tempo or 0),
        "pecasHora": float(r.pecas_hora or 0),
        "paradasMin": float(r.paradas_min or 0),
        "ciclo": float(r.ciclo or 0),
        "perdas": float(r.perdas or 0),
        "material": float(r.material or 0),
        "estampo": float(r.estampo or 0),
        "polimento": float(r.polimento or 0),
        "refilador": float(r.refilador or 0),
        "rebite": float(r.rebite or 0),
        "amassado": float(r.amassado or 0),
        "pintura": float(r.pintura or 0),
    }


@ensure_csrf_cookie
def controle_producao_cacarolas(request):
    nomes_produto = []
    nomes_produto.extend(
        Produto.objects.filter(ativo=True)
        .exclude(nome="")
        .values_list("nome", flat=True)
    )
    nomes_produto.extend(
        FiltroPedido.objects.exclude(descricao_produto="")
        .values_list("descricao_produto", flat=True)
    )
    vistos = set()
    produtos_seed = []
    for nome in nomes_produto:
        n = str(nome or "").strip()
        if not n:
            continue
        chave = n.casefold()
        if chave in vistos:
            continue
        vistos.add(chave)
        produtos_seed.append({"nome": n, "ciclo": 0})
        if len(produtos_seed) >= 5000:
            break
    contexto = {
        "produtos_seed_json": json.dumps(produtos_seed, ensure_ascii=False),
    }
    return render(request, "producao/controle_producao_cacarolas.html", contexto)


@ensure_csrf_cookie
def api_cacarolas_state(request):
    maquinas = list(CacarolaMaquina.objects.values_list("nome", flat=True))
    produtos = list(CacarolaProduto.objects.values("nome", "ciclo"))
    registros = [_serialize_cacarola_registro(r) for r in CacarolaRegistro.objects.all()[:5000]]
    return JsonResponse(
        {
            "maquinas": maquinas,
            "produtos": [{"nome": p["nome"], "ciclo": float(p["ciclo"] or 0)} for p in produtos],
            "registros": registros,
        }
    )


@require_POST
def api_cacarolas_add_maquina(request):
    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except Exception:  # noqa: BLE001
        payload = {}
    nome = str(payload.get("nome") or "").strip()
    if not nome:
        return JsonResponse({"ok": False, "erro": "Nome da máquina é obrigatório."}, status=400)
    CacarolaMaquina.objects.get_or_create(nome=nome)
    maquinas = list(CacarolaMaquina.objects.values_list("nome", flat=True))
    return JsonResponse({"ok": True, "maquinas": maquinas})


@require_POST
def api_cacarolas_update_maquina(request):
    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except Exception:  # noqa: BLE001
        payload = {}
    nome_atual = str(payload.get("nome_atual") or "").strip()
    novo_nome = str(payload.get("novo_nome") or "").strip()
    if not nome_atual or not novo_nome:
        return JsonResponse({"ok": False, "erro": "Nome atual e novo nome são obrigatórios."}, status=400)
    if nome_atual == novo_nome:
        maquinas = list(CacarolaMaquina.objects.values_list("nome", flat=True))
        return JsonResponse({"ok": True, "maquinas": maquinas})

    maq = CacarolaMaquina.objects.filter(nome=nome_atual).first()
    if not maq:
        return JsonResponse({"ok": False, "erro": "Máquina não encontrada."}, status=404)
    if CacarolaMaquina.objects.filter(nome=novo_nome).exists():
        return JsonResponse({"ok": False, "erro": "Já existe uma máquina com esse nome."}, status=400)

    maq.nome = novo_nome
    maq.save(update_fields=["nome"])
    CacarolaRegistro.objects.filter(maquina=nome_atual).update(maquina=novo_nome)
    maquinas = list(CacarolaMaquina.objects.values_list("nome", flat=True))
    return JsonResponse({"ok": True, "maquinas": maquinas})


@require_POST
def api_cacarolas_delete_maquina(request):
    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except Exception:  # noqa: BLE001
        payload = {}
    nome = str(payload.get("nome") or "").strip()
    senha = str(payload.get("senha") or "").strip()
    senha_ok = str(os.environ.get("CACAROLAS_DELETE_PASSWORD", "2568")).strip()
    if senha != senha_ok:
        return JsonResponse({"ok": False, "erro": "Senha inválida para excluir máquina."}, status=403)
    if not nome:
        return JsonResponse({"ok": False, "erro": "Nome da máquina é obrigatório."}, status=400)
    CacarolaMaquina.objects.filter(nome=nome).delete()
    maquinas = list(CacarolaMaquina.objects.values_list("nome", flat=True))
    return JsonResponse({"ok": True, "maquinas": maquinas})


@require_POST
def api_cacarolas_upsert_produtos(request):
    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except Exception:  # noqa: BLE001
        payload = {}
    produtos = payload.get("produtos") if isinstance(payload.get("produtos"), list) else []
    if not produtos:
        return JsonResponse({"ok": False, "erro": "Lista de produtos inválida."}, status=400)
    for p in produtos:
        nome = str((p or {}).get("nome") or "").strip()
        if not nome:
            continue
        ciclo = _dec_or_zero((p or {}).get("ciclo"))
        CacarolaProduto.objects.update_or_create(
            nome=nome,
            defaults={"ciclo": ciclo},
        )
    final = list(CacarolaProduto.objects.values("nome", "ciclo"))
    return JsonResponse({"ok": True, "produtos": [{"nome": x["nome"], "ciclo": float(x["ciclo"] or 0)} for x in final]})


@require_POST
def api_cacarolas_upsert_registro(request):
    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except Exception:  # noqa: BLE001
        payload = {}
    uid = str(payload.get("uid") or "").strip()
    data_txt = str(payload.get("data") or "").strip()
    if not uid or not data_txt:
        return JsonResponse({"ok": False, "erro": "UID e data são obrigatórios."}, status=400)
    try:
        data_val = datetime.strptime(data_txt, "%Y-%m-%d").date()
    except ValueError:
        return JsonResponse({"ok": False, "erro": "Data inválida."}, status=400)
    obj, _ = CacarolaRegistro.objects.update_or_create(
        uid=uid,
        defaults={
            "data": data_val,
            "turno": str(payload.get("turno") or "").strip(),
            "maquina": str(payload.get("maquina") or "").strip(),
            "responsavel": str(payload.get("responsavel") or "").strip(),
            "tipo_produto": str(payload.get("tipoProduto") or "").strip(),
            "produto": str(payload.get("produto") or "").strip(),
            "quantidade": _dec_or_zero(payload.get("quantidade")),
            "odf": str(payload.get("odf") or "").strip(),
            "inicio": _time_or_none(payload.get("inicio")),
            "fim": _time_or_none(payload.get("fim")),
            "refeicao": _dec_or_zero(payload.get("refeicao")),
            "tempo": _dec_or_zero(payload.get("tempo")),
            "pecas_hora": _dec_or_zero(payload.get("pecasHora")),
            "paradas_min": _dec_or_zero(payload.get("paradasMin")),
            "ciclo": _dec_or_zero(payload.get("ciclo")),
            "perdas": _dec_or_zero(payload.get("perdas")),
            "material": _dec_or_zero(payload.get("material")),
            "estampo": _dec_or_zero(payload.get("estampo")),
            "polimento": _dec_or_zero(payload.get("polimento")),
            "refilador": _dec_or_zero(payload.get("refilador")),
            "rebite": _dec_or_zero(payload.get("rebite")),
            "amassado": _dec_or_zero(payload.get("amassado")),
            "pintura": _dec_or_zero(payload.get("pintura")),
        },
    )
    return JsonResponse({"ok": True, "registro": _serialize_cacarola_registro(obj)})


@require_POST
def api_cacarolas_delete_registro(request):
    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except Exception:  # noqa: BLE001
        payload = {}
    uid = str(payload.get("uid") or "").strip()
    if not uid:
        return JsonResponse({"ok": False, "erro": "UID é obrigatório."}, status=400)
    CacarolaRegistro.objects.filter(uid=uid).delete()
    return JsonResponse({"ok": True})


def lista_produtos(request):
    produtos = Produto.objects.all().order_by("nome")
    return render(request, "producao/lista_produtos.html", {"produtos": produtos})


def novo_produto(request):
    if request.method == "POST":
        nome = request.POST.get("nome", "").strip()
        codigo = request.POST.get("codigo", "").strip()
        descricao = request.POST.get("descricao", "").strip()
        ativo = request.POST.get("ativo") == "on"

        if nome and codigo:
            Produto.objects.create(
                nome=nome,
                codigo=codigo,
                descricao=descricao,
                ativo=ativo,
            )
            return redirect(reverse("lista_produtos"))

    return render(request, "producao/novo_produto.html")


def lista_ordens(request):
    ordens = OrdemProducao.objects.select_related("produto").order_by("-data_criacao")
    return render(request, "producao/lista_ordens.html", {"ordens": ordens})


def nova_ordem(request):
    produtos = Produto.objects.filter(ativo=True).order_by("nome")

    if request.method == "POST":
        numero = request.POST.get("numero", "").strip()
        produto_id = request.POST.get("produto")
        quantidade = request.POST.get("quantidade")
        data_prevista = request.POST.get("data_prevista") or None
        observacoes = request.POST.get("observacoes", "").strip()

        if numero and produto_id and quantidade:
            produto = get_object_or_404(Produto, pk=produto_id)
            OrdemProducao.objects.create(
                numero=numero,
                produto=produto,
                quantidade=quantidade,
                data_prevista=data_prevista,
                observacoes=observacoes,
            )
            return redirect(reverse("lista_ordens"))

    contexto = {"produtos": produtos}
    return render(request, "producao/nova_ordem.html", contexto)


def atualizar_status_ordem(request, pk):
    ordem = get_object_or_404(OrdemProducao, pk=pk)
    if request.method == "POST":
        novo_status = request.POST.get("status")
        if novo_status in dict(OrdemProducao.STATUS_CHOICES):
            ordem.status = novo_status
            ordem.save()
        return redirect(reverse("lista_ordens"))

    contexto = {
        "ordem": ordem,
        "status_choices": OrdemProducao.STATUS_CHOICES,
    }
    return render(request, "producao/atualizar_status_ordem.html", contexto)
